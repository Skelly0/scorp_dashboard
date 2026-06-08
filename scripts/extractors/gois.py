"""Extract data for the GoIs page.

Reads from the live workbook's slimmed Politics sheet (rows 4-11 = 8 GoI slots,
currently 5 live + 3 reserved blanks). Per-row data is read by column offset
since the backend doesn't expose per-field named ranges. Sub-faction data is
zipped from the separate `SubFaction*` named ranges on the Sub-Factions sheet.
"""
from __future__ import annotations

import logging
import re
from typing import Any

from openpyxl.utils.cell import range_boundaries

from extractors._common import coerce_number, filter_blank_rows, read_named_range

WORLDVIEW_AXES = ["expansion", "authority", "corporate", "technocratic", "faith", "materialist"]

_log = logging.getLogger(__name__)

BLANK_TEXT_SENTINELS = {"", "none", "null"}

# GoI block lives at Politics rows 4-11 (8 slots, currently 5 live + 3 reserved blank).
GOI_BLOCK_FIRST_ROW = 4
GOI_BLOCK_LAST_ROW = 11

# Column offsets within the Politics GoI block.
COL_NAME = 1            # A
COL_INFLUENCE = 2       # B  (GM Override)
COL_UNITY = 3           # C
COL_EFFECTIVE_STANCE = 4  # D
COL_VOTE_WEIGHT = 5     # E
COL_APPROVAL = 6        # F
COL_ACTIVE_BENEFITS = 8  # H  (text "0 / 3 benefits unlocked")
COL_WORLDVIEW_FIRST = 11  # K..P  (Expn, Auth, Corp, Tech, Faith, Mat)
COL_MAD_INDEX = 17      # Q
COL_APPROACH = 18       # R
COL_DERIVED_INFLUENCE = 19  # S


def extract(wb) -> dict[str, Any]:
    classes = filter_blank_rows(read_named_range(wb, "ClassTable"))
    base_capture = read_named_range(wb, "PopCaptureBase")
    base_capture_headers = _headers_above_named_range(wb, "PopCaptureBase")
    captured_pop = read_named_range(wb, "GoIValueCapturedPop")
    captured_pop_headers = _headers_above_named_range(wb, "GoIValueCapturedPop")
    benefits_table = _read_goi_benefits_table(wb)

    # Read the GoI block directly from Politics by column offset.
    rows = _read_politics_goi_rows(wb)

    # Determine live indices (rows where col A name resolves to a non-blank string).
    live_indices = [i for i, r in enumerate(rows) if r["name"] not in (None, "")]
    live_names = [rows[i]["name"] for i in live_indices]

    main_classes = _infer_main_classes(
        live_names, base_capture, classes, base_capture_headers
    )

    sub_factions_by_goi = _sub_factions_by_goi(wb)

    out_gois: list[dict[str, Any]] = []
    for src_idx in live_indices:
        r = rows[src_idx]
        out_gois.append({
            "name": r["name"],
            "main_class": main_classes.get(r["name"]),
            "derived_influence": r["derived_influence"],
            "approval": r["approval"],
            "approach": r["approach"],
            "mad_index": r["mad_index"],
            "effective_worldview": r["worldview"],
            "active_benefits": _parse_active_benefits(
                r["active_benefits_text"], r["name"], benefits_table
            ),
            "sub_factions": sub_factions_by_goi.get(r["name"], []),
        })

    return {
        "gois": out_gois,
        "pop_capture_matrix": {
            "classes": [c[0] for c in classes],
            "gois": live_names,
            "values": _capture_values(
                wb,
                captured_pop,
                [c[0] for c in classes],
                live_names,
                live_indices,
                captured_pop_headers,
            ),
        },
    }


def _read_politics_goi_rows(wb):
    """Read Politics GoI block rows 4-11 by column offset."""
    if "Politics" not in wb.sheetnames:
        return []
    ws = wb["Politics"]
    out = []
    for row_num in range(GOI_BLOCK_FIRST_ROW, GOI_BLOCK_LAST_ROW + 1):
        worldview = {
            axis: coerce_number(ws.cell(row=row_num, column=COL_WORLDVIEW_FIRST + i).value)
            for i, axis in enumerate(WORLDVIEW_AXES)
        }
        out.append({
            "name": _clean_text(ws.cell(row=row_num, column=COL_NAME).value),
            "approval": coerce_number(ws.cell(row=row_num, column=COL_APPROVAL).value),
            "approach": _clean_text(ws.cell(row=row_num, column=COL_APPROACH).value),
            "mad_index": coerce_number(ws.cell(row=row_num, column=COL_MAD_INDEX).value),
            "derived_influence": coerce_number(ws.cell(row=row_num, column=COL_DERIVED_INFLUENCE).value),
            "worldview": worldview,
            "active_benefits_text": ws.cell(row=row_num, column=COL_ACTIVE_BENEFITS).value,
        })
    return out


_BENEFIT_COUNT_RE = re.compile(r"(\d+)\s*/\s*(\d+)")


def _read_goi_benefits_table(wb):
    named = read_named_range(wb, "GoIBenefitsTable")
    visible = _read_visible_goi_benefits_table(wb)
    if visible and len(visible) > len(named):
        return visible
    return named or visible


def _read_visible_goi_benefits_table(wb):
    if "GoI Benefits" not in wb.sheetnames:
        return []
    ws = wb["GoI Benefits"]
    header_row = None
    for row in range(1, min(ws.max_row, 40) + 1):
        first = ws.cell(row=row, column=1).value
        second = ws.cell(row=row, column=2).value
        third = ws.cell(row=row, column=3).value
        fourth = ws.cell(row=row, column=4).value
        if (
            first == "GoI"
            and second == "Threshold"
            and third == "Benefit Name"
            and fourth == "Description"
        ):
            header_row = row
            break
    if header_row is None:
        return []

    rows = []
    blank_streak = 0
    for row in range(header_row + 1, ws.max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, 5)]
        if all(v in (None, "") for v in values):
            if not rows:
                continue
            # Tolerate an isolated separator row between GoI groups; only a
            # larger gap signals the real end of the table. Breaking on the
            # first blank row truncated the table whenever the live sheet put
            # a spacer between groups — defeating the whole point of this
            # fallback (CLAUDE.md #51).
            blank_streak += 1
            if blank_streak >= 2:
                break
            continue
        blank_streak = 0
        rows.append(values)
    return rows


def _benefit_threshold_key(row):
    """Sort key: ascending threshold, with missing thresholds sorted last."""
    val = coerce_number(row[1]) if len(row) > 1 else None
    return (val is None, val if val is not None else 0.0)


def _parse_active_benefits(text, goi_name, benefits_table):
    matches = [r for r in benefits_table if r and r[0] == goi_name]
    # Benefits unlock in ascending-threshold order, so the workbook's
    # authoritative "N / M" count means the N lowest-threshold benefits are
    # active. Sort by threshold first so that rule holds regardless of the
    # row order the sheet happens to store them in.
    matches.sort(key=_benefit_threshold_key)
    m = _BENEFIT_COUNT_RE.search(str(text)) if text else None
    unlocked = int(m.group(1)) if m else 0
    total = int(m.group(2)) if m else len(matches)
    items = []
    for row in matches:
        name = row[2] if len(row) > 2 else None
        if name in (None, ""):
            continue
        description = row[3] if len(row) > 3 else None
        if description == "":
            description = None
        items.append({
            "name": name,
            "description": description,
            "threshold": coerce_number(row[1]) if len(row) > 1 else None,
            # Position among kept items (not raw row index) so a skipped
            # blank-name row can't shift the active/inactive boundary.
            "active": len(items) < unlocked,
        })
    unlocked_list = [b["name"] for b in items if b["active"]]
    if total == 0 and items:
        total = len(items)
    return {
        "unlocked": unlocked,
        "total": total,
        "unlocked_list": unlocked_list,
        "items": items,
    }


def _stance_row_to_worldview(row):
    """Map a 6-col SubFactionStances row to a {axis: value | None} dict.

    Returns None when the row is empty/all-None — the frontend treats that
    as "no stance" and hides the radar (graceful degradation when the GM
    hasn't filled the row yet, or the SubFactionStances range is missing).
    """
    if not row:
        return None
    out: dict[str, float | None] = {}
    for i, axis in enumerate(WORLDVIEW_AXES):
        out[axis] = coerce_number(row[i]) if i < len(row) else None
    if all(v is None for v in out.values()):
        return None
    return out


def _clean_text(value):
    if value in (None, ""):
        return None
    text = str(value).strip()
    if text.lower() in BLANK_TEXT_SENTINELS:
        return None
    return text


def _merge_sub_faction_record(base, overlay):
    if base is None:
        return overlay
    merged = dict(base)
    for key, value in overlay.items():
        if key == "minor_goals":
            if value:
                merged[key] = value
        elif key == "effective_worldview":
            if value is not None:
                merged[key] = value
        elif value not in (None, ""):
            merged[key] = value
    return merged


def _sub_factions_by_goi(wb):
    named = _named_sub_factions_by_goi(wb)
    visible = _visible_sub_factions_by_goi(wb)
    if not visible:
        return named

    out = {
        goi_name: [dict(record) for record in records]
        for goi_name, records in named.items()
    }
    for goi_name, records in visible.items():
        named_by_name = {
            record["name"]: record
            for record in named.get(goi_name, [])
            if record.get("name") not in (None, "")
        }
        out[goi_name] = [
            _merge_sub_faction_record(named_by_name.get(record["name"]), record)
            for record in records
        ]
    return out


def _named_sub_factions_by_goi(wb):
    """Zip the SubFaction* named ranges into per-GoI lists.

    Sub-Factions sheet layout (live wb, all ranges row-aligned at rows 6-20):
      A: GoI, B: SF name, C: Goal Axis, D: Goal Δ, E: Goal text,
      F: Influence, G-I: Minor Goals, J: Approval, L: National Share,
      N-S: 6-axis effective stance (Expn, Auth, Corp, Tech, Faith, Mat).

    The per-axis effective worldview is read directly from SubFactionStances
    (row-aligned with SubFactionGoals). Missing or empty rows yield None,
    and the panel hides the radar in that case.
    """
    goals = read_named_range(wb, "SubFactionGoals")
    influences = read_named_range(wb, "SubFactionInfluences")
    minor_goals = read_named_range(wb, "SubFactionMinorGoals")
    approvals = read_named_range(wb, "SubFactionApprovals")
    goals_text = read_named_range(wb, "SubFactionGoal")
    national_shares = read_named_range(wb, "SubFactionNationalShare")
    stances = read_named_range(wb, "SubFactionStances")

    by_goi: dict[str, list[dict[str, Any]]] = {}
    for i, gr in enumerate(goals):
        if not gr:
            continue
        goi_name = _clean_text(gr[0])
        sf_name = _clean_text(gr[1]) if len(gr) > 1 else None
        if goi_name is None or sf_name is None:
            continue
        infl = (
            coerce_number(influences[i][0])
            if i < len(influences) and influences[i]
            else None
        )
        appr = (
            coerce_number(approvals[i][0])
            if i < len(approvals) and approvals[i]
            else None
        )
        mgs = []
        if i < len(minor_goals) and minor_goals[i]:
            mgs = [g for g in (_clean_text(g) for g in minor_goals[i]) if g]
        goal_text = _clean_text(
            goals_text[i][0]
            if i < len(goals_text) and goals_text[i]
            else None
        )
        nat_share = (
            coerce_number(national_shares[i][0])
            if i < len(national_shares) and national_shares[i]
            else None
        )
        worldview = _stance_row_to_worldview(
            stances[i] if i < len(stances) else None
        )

        by_goi.setdefault(goi_name, []).append({
            "name": sf_name,
            "influence": infl,
            "approval": appr,
            "minor_goals": mgs,
            "goal": goal_text,
            "national_share": nat_share,
            "effective_worldview": worldview,
        })

    return by_goi


HEADER_ALIASES = {
    "goi": ("goi",),
    "sub_faction": ("sub-faction", "sub faction", "subfaction"),
    "goal": ("goal",),
    "influence": ("influence",),
    "approval": ("approval",),
    "national_share": ("national share",),
    "minor_1": ("minor 1", "minor goal 1"),
    "minor_2": ("minor 2", "minor goal 2"),
    "minor_3": ("minor 3", "minor goal 3"),
}

WORLDVIEW_HEADER_ALIASES = {
    "expansion": ("expn", "ideal expn", "expansion"),
    "authority": ("auth", "ideal auth", "authority"),
    "corporate": ("corp", "ideal corp", "corporate"),
    "technocratic": ("tech", "ideal tech", "technocratic"),
    "faith": ("faith", "ideal faith"),
    "materialist": ("mat", "ideal mat", "materialist"),
}


def _header_map(row):
    return {
        header: i
        for i, raw in enumerate(row)
        if (header := _clean_text(raw).lower() if _clean_text(raw) else None)
    }


def _first_header(headers, aliases):
    for alias in aliases:
        if alias in headers:
            return headers[alias]
    return None


def _row_value(row, idx):
    return row[idx] if idx is not None and idx < len(row) else None


def _visible_sub_factions_by_goi(wb):
    by_goi: dict[str, list[dict[str, Any]]] = {}
    for sheet_name in ("Modular Sub-Factions", "Sub-Factions", "Sub-Faction Detail"):
        source = _visible_sub_factions_from_sheet(wb, sheet_name)
        for goi_name, records in source.items():
            existing_by_name = {
                record["name"]: record
                for record in by_goi.get(goi_name, [])
                if record.get("name") not in (None, "")
            }
            by_goi[goi_name] = [
                _merge_sub_faction_record(existing_by_name.get(record["name"]), record)
                for record in records
            ]
    return by_goi


def _visible_sub_factions_from_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        return {}
    ws = wb[sheet_name]
    header_row = None
    headers = None
    for row_num in range(1, min(ws.max_row, 40) + 1):
        values = [ws.cell(row=row_num, column=col).value for col in range(1, ws.max_column + 1)]
        candidate_headers = _header_map(values)
        if (
            _first_header(candidate_headers, HEADER_ALIASES["goi"]) is not None
            and _first_header(candidate_headers, HEADER_ALIASES["sub_faction"]) is not None
        ):
            header_row = row_num
            headers = candidate_headers
            break
    if header_row is None or headers is None:
        return {}

    idx_goi = _first_header(headers, HEADER_ALIASES["goi"])
    idx_sub_faction = _first_header(headers, HEADER_ALIASES["sub_faction"])
    idx_goal = _first_header(headers, HEADER_ALIASES["goal"])
    idx_influence = _first_header(headers, HEADER_ALIASES["influence"])
    idx_approval = _first_header(headers, HEADER_ALIASES["approval"])
    idx_national_share = _first_header(headers, HEADER_ALIASES["national_share"])
    minor_idxs = [
        _first_header(headers, HEADER_ALIASES["minor_1"]),
        _first_header(headers, HEADER_ALIASES["minor_2"]),
        _first_header(headers, HEADER_ALIASES["minor_3"]),
    ]
    axis_idxs = {
        axis: _first_header(headers, aliases)
        for axis, aliases in WORLDVIEW_HEADER_ALIASES.items()
    }

    by_goi: dict[str, list[dict[str, Any]]] = {}
    blank_identity_streak = 0
    for row in ws.iter_rows(
        min_row=header_row + 1,
        max_row=ws.max_row,
        min_col=1,
        max_col=ws.max_column,
        values_only=True,
    ):
        goi_name = _clean_text(_row_value(row, idx_goi))
        sf_name = _clean_text(_row_value(row, idx_sub_faction))
        if goi_name is None and sf_name is None:
            if by_goi:
                blank_identity_streak += 1
                if blank_identity_streak >= 10:
                    break
            continue
        blank_identity_streak = 0
        if goi_name is None or sf_name is None:
            continue

        worldview = {
            axis: coerce_number(_row_value(row, idx))
            for axis, idx in axis_idxs.items()
        }
        if all(value is None for value in worldview.values()):
            worldview = None

        minor_goals = [
            goal
            for goal in (_clean_text(_row_value(row, idx)) for idx in minor_idxs)
            if goal
        ]
        record = {
            "name": sf_name,
            "influence": coerce_number(_row_value(row, idx_influence)),
            "approval": coerce_number(_row_value(row, idx_approval)),
            "minor_goals": minor_goals,
            "goal": _clean_text(_row_value(row, idx_goal)),
            "national_share": coerce_number(_row_value(row, idx_national_share)),
            "effective_worldview": worldview,
        }
        by_goi.setdefault(goi_name, []).append(record)
    return by_goi


def _headers_above_named_range(wb, range_name):
    """Return column headers from the row directly above a named data range.

    The live GoI matrices expose named ranges for the body only, while their
    header rows carry the GoI names. Reading those headers lets the extractor
    survive a GoI insertion/reorder without sliding values between columns.
    """
    if range_name not in wb.defined_names:
        return []
    dn = wb.defined_names[range_name]
    for sheet_name, ref in dn.destinations:
        clean_ref = ref.replace("$", "")
        min_col, min_row, max_col, _max_row = range_boundaries(clean_ref)
        header_row = min_row - 1
        if header_row < 1 or sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        return [
            ws.cell(row=header_row, column=col).value
            for col in range(min_col, max_col + 1)
        ]
    return []


def _column_index_by_header(headers):
    return {
        str(header): i
        for i, header in enumerate(headers)
        if header not in (None, "")
    }


def _infer_main_classes(live_names, capture_matrix, classes, capture_headers=None):
    """Best-effort: pick the class with the highest base capture for each GoI."""
    if not capture_matrix or not classes:
        return {}
    header_idx = _column_index_by_header(capture_headers or [])
    out = {}
    for j, name in enumerate(live_names):
        col_idx = header_idx.get(name, j)
        if col_idx >= len(capture_matrix[0] if capture_matrix else []):
            out[name] = None
            continue
        best_i = max(
            range(len(capture_matrix)),
            key=lambda i: coerce_number(capture_matrix[i][col_idx]) or 0,
        )
        out[name] = classes[best_i][0] if best_i < len(classes) else None
    return out


def _find_cell(ws, value):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value == value:
                return cell.row, cell.column
    return None


def _capture_values_from_titled_block(wb, class_names, live_names):
    """Read the displayed GOI VALUE CAPTURED POP table by labels.

    The live workbook's visible table grew to include Security before the
    backing named range did. Label-based extraction keeps the dashboard aligned
    with the player-visible source even when workbook names lag by a column.
    """
    if "Party and GoI Pop Capture" not in wb.sheetnames:
        return None

    ws = wb["Party and GoI Pop Capture"]
    anchor = _find_cell(ws, "GOI VALUE CAPTURED POP")
    if anchor is None:
        return None

    title_row, title_col = anchor
    header_row = title_row + 2
    goi_cols = {}
    in_header_block = False
    for col in range(title_col + 1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        if header in (None, ""):
            if in_header_block:
                break
            continue
        in_header_block = True
        if header in live_names:
            goi_cols[header] = col

    if not all(name in goi_cols for name in live_names):
        return None

    rows_by_class = {}
    for row in range(header_row + 1, ws.max_row + 1):
        class_name = ws.cell(row=row, column=title_col).value
        if class_name in (None, "") and rows_by_class:
            break
        if class_name in class_names:
            rows_by_class[class_name] = [
                coerce_number(ws.cell(row=row, column=goi_cols[name]).value)
                for name in live_names
            ]
        if len(rows_by_class) == len(class_names):
            break

    return [
        rows_by_class.get(class_name, [None] * len(live_names))
        for class_name in class_names
    ]


def _capture_values(wb, capture, class_names, live_names, live_indices, capture_headers=None):
    titled_values = _capture_values_from_titled_block(wb, class_names, live_names)
    if titled_values is not None:
        return titled_values

    header_idx = _column_index_by_header(capture_headers or [])
    n_classes = len(class_names)
    rows = []
    for i in range(n_classes):
        if i >= len(capture):
            rows.append([None] * len(live_names))
            continue
        row = []
        for fallback_idx, name in zip(live_indices, live_names, strict=False):
            col_idx = header_idx.get(name, fallback_idx)
            row.append(
                coerce_number(capture[i][col_idx])
                if col_idx < len(capture[i])
                else None
            )
        rows.append(row)
    return rows
