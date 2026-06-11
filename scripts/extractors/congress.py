"""Extract the All-Worker Congress page data.

Two views of the same 27-seat chamber, both on the `All-Worker Congress`
sheet and sharing one party-column axis (15 party slots in cols B..P
mirroring the Parties sheet, plus Non-aligned in col Q):

- Party totals: seats per party across the whole Congress.
- Trade Federation delegations: the `DELEGATION → PARTY SEATS` matrix
  (Art. 15) — one row per federation, the largest-remainder split of each
  delegation across parties.

Matrix source preference (the gotcha-#51 named-range-then-visible-table
pattern):

1. `CongressDelegationSeats` named range when the GM has pinned it: col A
   federation names + the same party columns as `CongressPartyNames`, data
   rows only (no title/header rows).
2. Title-located fallback: the block whose col-A title contains
   `DELEGATION` + `PARTY SEATS` (the sibling `... PARTY QUOTAS` helper
   block is skipped) on the sheet hosting `CongressPartyNames`. Data rows
   start two rows below the title (title row, then `Federation \\ Party`
   header row) and stop at the first blank col-A cell.

Congress party totals are DERIVED from matrix column sums whenever the
matrix is readable — the live workbook's `CongressPartySeats` row has
drifted to literal zeros while the matrix kept computing, so the matrix is
authoritative. The `CongressPartySeats` row remains the fallback when no
matrix is found.

Everything here is soft-optional (validate_schema.SOFT_OPTIONAL_V3_RANGES):
a workbook without the ranges yields empty party/delegation lists and the
frontend renders an empty state. Named zero-seat parties are kept at the
chamber level (a founded party shut out of the chamber is information);
blank-name slots are dropped (convention 8). Zero-seat parties are omitted
WITHIN each delegation — at that level they are noise, not information.

The Celestial Council allocation (`CouncilSeatsByParty`) is no longer
extracted; the dashboard's second band shows the federation delegations.
"""
from __future__ import annotations

from typing import Any

from openpyxl.utils import range_boundaries

from extractors._common import coerce_number, read_named_range

# Matrix rows whose col-A label matches these are structure, not federations.
_NON_FEDERATION_LABELS = ("TOTAL", "FEDERATION")
_MAX_FEDERATION_ROWS = 40


def extract(wb) -> dict[str, Any]:
    names = [_clean(v) for v in _single_row(read_named_range(wb, "CongressPartyNames"))]
    delegations = _read_delegations(wb, names)
    if delegations:
        chamber = _chamber_from_delegations(names, delegations)
    else:
        chamber = _chamber(names, _single_row(read_named_range(wb, "CongressPartySeats")))
    return {
        "congress": chamber,
        "federations": {
            "total_seats": int(sum(d["seats"] for d in delegations)),
            "delegations": delegations,
        },
    }


def _single_row(rows: list[list[Any]]) -> list[Any]:
    return rows[0] if rows else []


def _clean(value: Any) -> str | None:
    if value in (None, ""):
        return None
    name = str(value).strip()
    return name or None


def _chamber(names: list[str | None], seats: list[Any]) -> dict[str, Any]:
    parties: list[dict[str, Any]] = []
    total = 0.0
    for i, name in enumerate(names):
        if name is None:
            continue  # blank party slot reserved for future growth
        seat_count = coerce_number(seats[i]) if i < len(seats) else None
        parties.append({"name": name, "seats": seat_count})
        if seat_count is not None:
            total += seat_count
    return {"total_seats": int(total), "parties": parties}


def _chamber_from_delegations(
    names: list[str | None], delegations: list[dict[str, Any]]
) -> dict[str, Any]:
    totals: dict[str, float] = {}
    for d in delegations:
        for p in d["parties"]:
            totals[p["name"]] = totals.get(p["name"], 0.0) + p["seats"]
    parties = [{"name": n, "seats": totals.get(n, 0.0)} for n in names if n is not None]
    return {"total_seats": int(sum(totals.values())), "parties": parties}


def _read_delegations(wb, names: list[str | None]) -> list[dict[str, Any]]:
    if not any(n is not None for n in names):
        return []  # no party axis — the matrix cells cannot be labelled
    rows = read_named_range(wb, "CongressDelegationSeats")
    if not rows:
        rows = _matrix_rows_by_title(wb)
    delegations: list[dict[str, Any]] = []
    for row in rows:
        fed = _clean(row[0]) if row else None
        if fed is None or any(fed.upper().startswith(lbl) for lbl in _NON_FEDERATION_LABELS):
            continue
        parties: list[dict[str, Any]] = []
        total = 0.0
        for i, name in enumerate(names):
            if name is None:
                continue
            seats = coerce_number(row[i + 1]) if i + 1 < len(row) else None
            if seats is None:
                continue
            total += seats
            if seats > 0:
                parties.append({"name": name, "seats": seats})
        delegations.append({"name": fed, "seats": int(total), "parties": parties})
    return delegations


def _party_columns(wb) -> tuple[str, int, int] | None:
    """Sheet + column span of the shared party axis (from CongressPartyNames)."""
    if "CongressPartyNames" not in wb.defined_names:
        return None
    for sheet_name, ref in wb.defined_names["CongressPartyNames"].destinations:
        min_col, _r1, max_col, _r2 = range_boundaries(ref)
        return sheet_name, min_col, max_col
    return None


def _matrix_rows_by_title(wb) -> list[list[Any]]:
    bounds = _party_columns(wb)
    if bounds is None:
        return []
    sheet_name, min_col, max_col = bounds
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]

    title_row = None
    for (cell,) in ws.iter_rows(min_col=1, max_col=1):
        text = str(cell.value).strip().upper() if isinstance(cell.value, str) else ""
        if "DELEGATION" in text and "PARTY SEATS" in text:
            title_row = cell.row
            break
    if title_row is None:
        return []

    rows: list[list[Any]] = []
    first_data_row = title_row + 2  # skip the `Federation \ Party` header row
    for r in range(first_data_row, first_data_row + _MAX_FEDERATION_ROWS):
        fed = ws.cell(row=r, column=1).value
        if fed in (None, ""):
            break
        rows.append([fed] + [ws.cell(row=r, column=c).value for c in range(min_col, max_col + 1)])
    return rows
