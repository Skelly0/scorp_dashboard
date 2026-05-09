"""Sync orchestrator: download xlsx, validate, extract per-page JSON, write meta.

Designed to run inside a GitHub Action; can also be run locally for testing.
"""
from __future__ import annotations

import argparse
import json
import logging
import os
import sys
import tempfile
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl
import requests

from extractors import catalog as ex_catalog
from extractors import cropsim as ex_cropsim
from extractors import demographics as ex_demographics
from extractors import gois as ex_gois
from extractors import map as ex_map
from extractors import parties as ex_parties
from extractors import population as ex_population
from extractors import pops as ex_pops
from extractors import senate as ex_senate
from extractors import situations as ex_situations
from extractors import status as ex_status
from extractors import tech as ex_tech
from history import write_snapshot as write_history_snapshot
from notify_telegram import build_message, send as send_telegram
from validate_schema import SchemaValidationError, validate

logger = logging.getLogger(__name__)

SCHEMA_VERSION = 9
EXPORT_URL = "https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"
RETRY_BACKOFFS = [5, 15, 45]  # seconds


@dataclass
class SyncResult:
    status: str
    partial_failures: list[str]


def download_xlsx(sheet_id: str, dest: Path) -> Path:
    url = EXPORT_URL.format(sheet_id=sheet_id)
    last_exc: Exception | None = None
    for i, delay in enumerate([0, *RETRY_BACKOFFS]):
        if delay:
            time.sleep(delay)
        try:
            r = requests.get(url, timeout=60, allow_redirects=True)
            r.raise_for_status()
            dest.write_bytes(r.content)
            return dest
        except Exception as exc:  # noqa: BLE001
            logger.warning("Download attempt %d failed: %s", i + 1, exc)
            last_exc = exc
    raise RuntimeError(f"Failed to download xlsx after {len(RETRY_BACKOFFS) + 1} attempts: {last_exc}")


def write_json_atomic(path: Path, data: dict | list) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    tmp.write_text(json.dumps(data, indent=2, sort_keys=True))
    tmp.replace(path)


def read_senate_flag(wb) -> bool:
    """Read Var_SenatePageVisible. Convention: single-cell named range (e.g. Variable!$B$1).

    Defensive: if someone defines it as a 1×1 range (e.g. $B$1:$B$1), openpyxl returns
    a tuple-of-tuples for `ws[ref]`, which is truthy regardless of value. We pull the
    top-left cell explicitly to handle both cases.
    """
    if "Var_SenatePageVisible" not in wb.defined_names:
        return False
    dn = wb.defined_names["Var_SenatePageVisible"]
    for sheet_name, ref in dn.destinations:
        clean = ref.replace("$", "")
        result = wb[sheet_name][clean]
        # ws[address] returns either a Cell (single addr) or tuple-of-tuples (range).
        if isinstance(result, tuple):
            cell = result[0][0]
        else:
            cell = result
        return bool(cell.value)
    return False


def _sheet_modified_time(wb) -> str | None:
    mod = wb.properties.modified
    if mod is None:
        return None
    return mod.replace(tzinfo=timezone.utc).isoformat() if mod.tzinfo is None else mod.isoformat()


def run_sync(xlsx_path: Path, out_dir: Path) -> SyncResult:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    senate_enabled = read_senate_flag(wb)
    validate(wb, senate_enabled=senate_enabled)

    partial_failures: list[str] = []
    status_data: dict[str, Any] | None = None

    extractors = [
        ("status", ex_status.extract),
        ("population", ex_population.extract),
        ("pops", ex_pops.extract),
        ("demographics", ex_demographics.extract),
        ("cropsim", ex_cropsim.extract),
        ("gois", ex_gois.extract),
        ("parties", ex_parties.extract),
        ("map", ex_map.extract),
        ("catalog", ex_catalog.extract),
        ("situations", ex_situations.extract),
        ("tech", ex_tech.extract),
        # Phase 2 extractors registered here as they land.
    ]

    for page_name, fn in extractors:
        try:
            data = fn(wb)
            write_json_atomic(out_dir / f"{page_name}.json", data)
            if page_name == "status":
                status_data = data
        except Exception as exc:  # noqa: BLE001 — keep going on per-page failure
            logger.error("Extractor %s failed: %s", page_name, exc)
            partial_failures.append(page_name)

    if senate_enabled:
        try:
            data = ex_senate.extract(wb)
            write_json_atomic(out_dir / "senate.json", data)
        except Exception as exc:  # noqa: BLE001
            logger.error("Senate extractor failed: %s", exc)
            partial_failures.append("senate")
    else:
        # Make sure stale senate.json from a prior ON state is removed.
        senate_path = out_dir / "senate.json"
        if senate_path.exists():
            senate_path.unlink()

    synced_at = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")

    history_year: int | None = None
    if status_data is not None:
        try:
            wrote = write_history_snapshot(out_dir, status_data.get("year"), status_data, synced_at)
            if wrote:
                history_year = status_data.get("year")
        except Exception as exc:  # noqa: BLE001 — history is best-effort, never blocks sync
            logger.error("History snapshot failed: %s", exc)

    meta = {
        "synced_at": synced_at,
        "sheet_modified_time": _sheet_modified_time(wb),
        "senate_visible": senate_enabled,
        "schema_version": SCHEMA_VERSION,
        "partial_failures": partial_failures,
        "history_year": history_year,
    }
    write_json_atomic(out_dir / "meta.json", meta)

    return SyncResult(status="ok", partial_failures=partial_failures)


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    parser = argparse.ArgumentParser()
    parser.add_argument("--out-dir", default="public/data", help="JSON output directory")
    parser.add_argument("--sheet-id", default=os.environ.get("SHEET_ID"))
    parser.add_argument("--xlsx", help="Use a local xlsx instead of downloading (for testing)")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    try:
        if args.xlsx:
            xlsx_path = Path(args.xlsx)
        else:
            if not args.sheet_id:
                raise RuntimeError("SHEET_ID env var or --sheet-id flag required")
            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                xlsx_path = download_xlsx(args.sheet_id, Path(tmp.name))
        result = run_sync(xlsx_path, out_dir)
        if result.partial_failures:
            logger.warning("Partial failures: %s", result.partial_failures)
        logger.info("Sync OK")
        return 0
    except (RuntimeError, SchemaValidationError) as exc:
        logger.exception("Sync failed")
        run_url = os.environ.get("GITHUB_RUN_URL", "(no run URL)")
        send_telegram(build_message(summary=str(exc), run_url=run_url))
        return 1


if __name__ == "__main__":
    sys.exit(main())
