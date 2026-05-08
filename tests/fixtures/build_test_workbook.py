"""Build a minimal test workbook that mirrors the live SCORP Colony schema.

Used as a fixture by extractor tests. Not committed — regenerated on each test run.
"""
from __future__ import annotations

from pathlib import Path

import openpyxl
from openpyxl.workbook.defined_name import DefinedName


def build(out_path: Path) -> Path:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # ---- Reference sheet ----
    ref = wb.create_sheet("Reference")
    ref["A1"] = "ClassTable"
    # ClassTable: Name, Tier, StartingPop, PoliticalWeight (cols A-D from row 56-70 in real wb;
    # we just need the shape — 15 rows, 11 live + 4 blank).
    classes = [
        ("Bureaucrats", "Upper", 970, 3.0),
        ("Capitalists", "Upper", 220, 5.0),
        ("Engineers", "Middle", 1115, 2.0),
        ("Scientists", "Middle", 1790, 2.5),
        ("Security", "Middle", 1040, 1.5),
        ("Proprietors", "Middle", 295, 1.5),
        ("Managerial", "Middle", 295, 2.5),
        ("Agricultural Workers", "Lower", 1115, 0.8),
        ("Industrial Workers", "Lower", 4170, 0.8),
        ("Service Workers", "Lower", 3870, 0.7),
        ("Skilled Tradesmen", "Lower", 990, 0.9),
    ]
    for i, (name, tier, pop, weight) in enumerate(classes, start=56):
        ref.cell(row=i, column=1, value=name)
        ref.cell(row=i, column=2, value=tier)
        ref.cell(row=i, column=3, value=pop)
        ref.cell(row=i, column=4, value=weight)
    # 4 blank reserved slots (rows 67-70 already empty by default).

    _add_name(wb, "ClassTable", "Reference!$A$56:$D$70")

    # ---- Variable sheet ----
    var = wb.create_sheet("Variable")
    var["A1"], var["B1"] = "Var_SenatePageVisible", False
    _add_name(wb, "Var_SenatePageVisible", "Variable!$B$1")
    # Year mirror (Colony!H1 in live wb; named range so extractors don't read by addr).
    var["A2"], var["B2"] = "Var_Year", 12
    _add_name(wb, "Var_Year", "Variable!$B$2")
    # v3 tuning knobs (soft-optional in validator; extractors check presence).
    var["A3"], var["B3"] = "Var_BaseDeathRate", 0.012
    _add_name(wb, "Var_BaseDeathRate", "Variable!$B$3")
    var["A4"], var["B4"] = "Var_HousingOvercrowdingExp", 1.5
    _add_name(wb, "Var_HousingOvercrowdingExp", "Variable!$B$4")
    var["A5"], var["B5"] = "Var_BaseGrowthRate", 0.020
    _add_name(wb, "Var_BaseGrowthRate", "Variable!$B$5")
    var["A6"], var["B6"] = "Var_GrowthSatElasticity", 0.95
    _add_name(wb, "Var_GrowthSatElasticity", "Variable!$B$6")

    # ---- Politics sheet ----
    pol = wb.create_sheet("Politics")
    pol["B1"] = 0.42      # Stability
    pol["E1"] = 0.38      # Crisis Factor
    _add_name(wb, "Stability", "Politics!$B$1")
    _add_name(wb, "CrisisFactor", "Politics!$E$1")
    # Overton window B13:B18
    overton = [5.0, 4.5, 5.0, 4.0, 4.0, 4.0]
    for i, v in enumerate(overton, start=13):
        pol.cell(row=i, column=2, value=v)
    _add_name(wb, "OvertonExpn", "Politics!$B$13")
    _add_name(wb, "OvertonAuth", "Politics!$B$14")
    _add_name(wb, "OvertonCorp", "Politics!$B$15")
    _add_name(wb, "OvertonTech", "Politics!$B$16")
    _add_name(wb, "OvertonFaith", "Politics!$B$17")
    _add_name(wb, "OvertonMat", "Politics!$B$18")
    pol["B2"] = 0.62  # Effective Gov Approval
    _add_name(wb, "EffectiveGovApproval", "Politics!$B$2")

    # ---- Colony sheet (Treasury + resources) ----
    col = wb.create_sheet("Colony")
    col["A1"], col["B1"] = "Money", 487
    col["A2"], col["B2"] = "Money_Delta", -12
    resources = [
        ("Food", 0, -2),
        ("Materials", 200, -4),
        ("Ore", 100, 0),
        ("Energy", 50, 0),
        ("Housing", -500, 0),
        ("He-3", 0, 1),
        ("Water", 60, -1),
    ]
    for i, (name, current, delta) in enumerate(resources, start=4):
        col.cell(row=i, column=1, value=name)
        col.cell(row=i, column=2, value=current)
        col.cell(row=i, column=3, value=delta)
    _add_name(wb, "TreasuryMoney", "Colony!$B$1")
    _add_name(wb, "TreasuryMoneyDelta", "Colony!$B$2")
    _add_name(wb, "ResourceFlows", "Colony!$A$4:$C$10")
    col["A12"], col["B12"], col["C12"], col["I12"] = "Housing", "capacity", 16500, 0.96
    _add_name(wb, "HousingCapacity", "Colony!$C$12")
    _add_name(wb, "HousingRatio", "Colony!$I$12")

    # ---- Popsim sheet ----
    pop = wb.create_sheet("Popsim")
    # Just enough to satisfy population total derivation. PopsimPop = B5:B19
    for i, (name, _, p, _w) in enumerate(classes, start=5):
        pop.cell(row=i, column=1, value=name)
        pop.cell(row=i, column=2, value=p)
    _add_name(wb, "PopsimPop", "Popsim!$B$5:$B$19")

    # Popsim worldview block (cols B-G, rows 41-55, paired with class names in col A).
    # We mirror class names from ClassTable into col A using a formula approximation —
    # for the fixture, just write them directly.
    for i, (name, _, _, _) in enumerate(classes, start=41):
        pop.cell(row=i, column=1, value=name)
        # Six axis values per class — make them deterministic but varied.
        for axis in range(6):
            pop.cell(row=i, column=2 + axis, value=4.0 + (i % 3) * 0.5 - axis * 0.3)
    _add_name(wb, "PopsimWorldview", "Popsim!$B$41:$G$55")

    # Popsim Wealth & Income block (rows 61-75: cols A name, B gross/cap, D income tax/cap,
    # E wealth tax/cap, F effective rate, G disposable/cap, I wealth/cap, J total class income).
    # Also Standard of Living rows 97-111, Social Privileges rows 79-93, Status rows 133-147,
    # Satisfaction M151:M165. We won't reproduce real backend formulas — just write
    # deterministic numbers so the extractor has something to read.
    for i, (name, _, p, _w) in enumerate(classes, start=61):
        pop.cell(row=i, column=1, value=name)
        gross = 12.0 + (i - 61)
        pop.cell(row=i, column=2, value=gross)
        pop.cell(row=i, column=4, value=gross * 0.10)  # income tax/cap
        pop.cell(row=i, column=5, value=gross * 0.02)  # wealth tax/cap
        pop.cell(row=i, column=6, value=0.12)          # effective rate
        pop.cell(row=i, column=7, value=gross * 0.88)  # disposable/cap
        pop.cell(row=i, column=9, value=gross * 4)     # wealth/cap
        pop.cell(row=i, column=10, value=gross * 0.88 * p)  # total class income (post-tax)
    _add_name(wb, "PopsimGrossPerCap", "Popsim!$B$61:$B$75")
    _add_name(wb, "PopsimDisposablePerCap", "Popsim!$G$61:$G$75")
    _add_name(wb, "PopsimWealthPerCap", "Popsim!$I$61:$I$75")
    # WealthIncomePerClass: full A:J × 15 rows. Extractor indexes within each row.
    # Live workbook puts this at A62:J76; fixture data lives one row earlier.
    _add_name(wb, "WealthIncomePerClass", "Popsim!$A$61:$J$75")

    # WorkforceSupplyDemand: A name, B supply, C demand, D fill ratio, E unemployment.
    # 14 rows in live workbook; seed deterministic data for first 11 classes.
    for i, (name, _, p, _w) in enumerate(classes, start=24):
        pop.cell(row=i, column=1, value=name)
        supply = float(p)
        demand = float(p) * 1.05
        pop.cell(row=i, column=2, value=supply)
        pop.cell(row=i, column=3, value=demand)
        pop.cell(row=i, column=4, value=min(supply / demand, 1.0))
        pop.cell(row=i, column=5, value=0.05 + (i - 24) * 0.01)
    _add_name(wb, "WorkforceSupplyDemand", "Popsim!$A$24:$E$37")

    # PopsimUnemployed: raw count (separate from WorkforceSupplyDemand col E ratio).
    # Live wb places at E25:E39; fixture uses col K to avoid overlap.
    for i, (name, _, p, _w) in enumerate(classes, start=24):
        pop.cell(row=i, column=11, value=int(p * (0.05 + (i - 24) * 0.005)))
    _add_name(wb, "PopsimUnemployed", "Popsim!$K$24:$K$37")
    # Mobility In/Out — fixture only; live workbook puts these on Popsim cols F/G
    # (same table as growth/death). Names are soft-optional in validate_schema so a
    # workbook without them still syncs; the frontend renders "—".
    for i, (_, _, _, _) in enumerate(classes, start=24):
        pop.cell(row=i, column=12, value=2 + (i - 24))   # L: mobility in
        pop.cell(row=i, column=13, value=1 + (i - 24))   # M: mobility out
    _add_name(wb, "PopsimMobilityIn", "Popsim!$L$24:$L$37")
    _add_name(wb, "PopsimMobilityOut", "Popsim!$M$24:$M$37")

    # Standard of Living: rows 97-111, cols B (SoL), C (Expected SoL).
    for i, _ in enumerate(classes, start=97):
        pop.cell(row=i, column=1, value=classes[i - 97][0])
        pop.cell(row=i, column=2, value=0.42 + (i - 97) * 0.01)
        pop.cell(row=i, column=3, value=0.55)
    _add_name(wb, "PopsimSoL", "Popsim!$B$97:$B$111")
    _add_name(wb, "PopsimExpectedSoL", "Popsim!$C$97:$C$111")

    # Social Privileges: rows 79-93, col B.
    for i, _ in enumerate(classes, start=79):
        pop.cell(row=i, column=1, value=classes[i - 79][0])
        pop.cell(row=i, column=2, value=0.30 + (i - 79) * 0.02)
    _add_name(wb, "PopsimSocialPrivileges", "Popsim!$B$79:$B$93")

    # Status block: rows 133-147. C radicalisation, E abject poverty, G organisation, H literacy,
    # J vote eligibility, L votes total, M vote share.
    for i, _ in enumerate(classes, start=133):
        pop.cell(row=i, column=1, value=classes[i - 133][0])
        pop.cell(row=i, column=3, value=0.20 + (i - 133) * 0.01)  # radicalisation
        pop.cell(row=i, column=5, value=0.10)                      # abject poverty
        pop.cell(row=i, column=7, value=0.30)                      # organisation
        pop.cell(row=i, column=8, value=0.65)                      # literacy
        pop.cell(row=i, column=10, value=0.85)                     # vote eligibility
        pop.cell(row=i, column=12, value=int(classes[i - 133][2] * 0.85))  # votes total
        pop.cell(row=i, column=13, value=0.10)                     # vote share
    _add_name(wb, "PopsimRadicalisation", "Popsim!$C$133:$C$147")
    _add_name(wb, "PopsimAbjectPoverty", "Popsim!$E$133:$E$147")
    _add_name(wb, "PopsimOrganisation", "Popsim!$G$133:$G$147")
    _add_name(wb, "PopsimLiteracy", "Popsim!$H$133:$H$147")
    _add_name(wb, "PopsimVotesTotal", "Popsim!$L$133:$L$147")
    _add_name(wb, "PopsimVoteShare", "Popsim!$M$133:$M$147")

    # Satisfaction M151:M165
    for i, _ in enumerate(classes, start=151):
        pop.cell(row=i, column=13, value=0.40)
    _add_name(wb, "PopsimSatisfaction", "Popsim!$M$151:$M$165")

    # ---- Wages & Welfare sheet ----
    ww = wb.create_sheet("Wages & Welfare")
    # AdditionalIncomeRange = H23:I37 — col H per-class total Additional Income, col I label.
    # For test purposes, we fake the breakdown across cols B-E for rows 23-37,
    # with the total in col F mirrored to H.
    for i, _ in enumerate(classes, start=23):
        ww.cell(row=i, column=1, value=classes[i - 23][0])
        ww.cell(row=i, column=2, value=0.5)   # welfare
        ww.cell(row=i, column=3, value=0.0)   # dividends
        ww.cell(row=i, column=4, value=0.2)   # subsidies
        ww.cell(row=i, column=5, value=0.0)   # other
        ww.cell(row=i, column=6, value=0.7)   # total
        ww.cell(row=i, column=8, value=0.7)   # mirror for AdditionalIncomeRange
        ww.cell(row=i, column=9, value="total")
    _add_name(wb, "AdditionalIncomeRange", "'Wages & Welfare'!$H$23:$I$37")
    # Per-class breakdown read directly (cols B-E) — track via an aux range too.
    _add_name(wb, "AdditionalIncomeBreakdown", "'Wages & Welfare'!$A$23:$F$37")

    # ---- Mortality sheet (v3) ----
    mort = wb.create_sheet("Mortality")
    # Per-class mortality rates and deaths/turn; rows 13-23 (11 classes).
    # Live workbook spans 13-27 (15 slots); fixture only fills 11.
    for i, (name, _, p, _w) in enumerate(classes, start=13):
        mort.cell(row=i, column=1, value=name)
        rate = 0.010 + (i - 13) * 0.001  # 0.010 .. 0.020
        mort.cell(row=i, column=6, value=rate)         # F mortality rate
        mort.cell(row=i, column=7, value=int(p * rate))  # G deaths/turn
        mort.cell(row=i, column=8, value=int(p * rate) + 4)  # H births/turn (deaths + 4 for variety)
    mort["A30"], mort["B30"] = "Total deaths/turn", 280  # round figure
    mort["A31"], mort["B31"] = "Effective CDR", 0.0125
    mort["A32"], mort["B32"] = "Total births/turn", 320  # round figure: net = +40
    _add_name(wb, "MortalityRates", "Mortality!$F$13:$F$27")
    _add_name(wb, "DeathsPerTurn", "Mortality!$G$13:$G$27")
    _add_name(wb, "ClassBirths", "Mortality!$H$13:$H$27")
    _add_name(wb, "TotalDeathsPerTurn", "Mortality!$B$30")
    _add_name(wb, "EffectiveCDR", "Mortality!$B$31")
    _add_name(wb, "TotalBirths", "Mortality!$B$32")

    # ---- Housing sheet (v3, just for HousingGrowthMult) ----
    hou = wb.create_sheet("Housing")
    hou["A12"], hou["B12"] = "Housing growth mult", 0.92
    _add_name(wb, "HousingGrowthMult", "Housing!$B$12")

    # ---- Cropsim sheet (v3 food security) ----
    cs = wb.create_sheet("Cropsim")
    cs["A26"], cs["B26"] = "Food security ratio", 1.05
    cs["A27"], cs["B27"] = "Food per cap", 1.20
    cs["A28"], cs["B28"] = "Food variety index", 0.78
    _add_name(wb, "FoodSecurityRatio", "Cropsim!$B$26")
    _add_name(wb, "FoodPerCap", "Cropsim!$B$27")
    _add_name(wb, "FoodVarietyIndex", "Cropsim!$B$28")

    # Politics GoI block: rows 4-11 (8 slots, 4 live + 4 blank)
    gois = [
        ("Founders", "Bureaucrats", 0.30, 0.55, "Reformist"),
        ("Capitalists", "Capitalists", 0.28, 0.50, "Pragmatic"),
        ("Security", "Security", 0.20, 0.60, "Defensive"),
        ("Unionists", "Industrial Workers", 0.22, 0.45, "Activist"),
    ]
    for i, (name, main_class, infl, appr, approach) in enumerate(gois, start=4):
        pol.cell(row=i, column=1, value=name)
        pol.cell(row=i, column=2, value=infl)        # GM-override influence
        pol.cell(row=i, column=6, value=appr)        # Approval (col F)
        pol.cell(row=i, column=9, value=int(8 * infl))  # Council seats (col I)
        # Effective worldview cols K-P (11-16)
        for axis in range(6):
            pol.cell(row=i, column=11 + axis, value=4.0 + (i - 4) * 0.2 + axis * 0.1)
        pol.cell(row=i, column=17, value=0.10 + (i - 4) * 0.05)  # Mad Index col Q
        pol.cell(row=i, column=18, value=approach)                # Approach col R
        pol.cell(row=i, column=19, value=infl)                    # Derived Influence col S
        # Active Benefits: extractor reads col H (column 8) by direct offset
        # per gois.py:COL_ACTIVE_BENEFITS. The col T cell + GoIActiveBenefits
        # named range are kept for documentation/consistency but unused by reads.
        pol.cell(row=i, column=8, value=f"{i - 3} / 3 unlocked")   # Active Benefits col H
        pol.cell(row=i, column=20, value=f"{i - 3} / 3 unlocked")  # mirror to col T
    _add_name(wb, "GoINames", "Politics!$A$4:$A$11")
    _add_name(wb, "GoIDerivedInfluence", "Politics!$S$4:$S$11")
    _add_name(wb, "GoIApproval", "Politics!$F$4:$F$11")
    _add_name(wb, "GoIEffectiveWorldview", "Politics!$K$4:$P$11")
    _add_name(wb, "GoIMadIndex", "Politics!$Q$4:$Q$11")
    _add_name(wb, "GoIApproach", "Politics!$R$4:$R$11")
    _add_name(wb, "GoIActiveBenefits", "Politics!$T$4:$T$11")

    # Sub-faction block — mirrors live wb's `Sub-Factions` sheet col-for-col
    # (A:L) so per-column named-range slices line up. Parked here on Politics
    # at cols U-AF (offset 21..32) instead of a dedicated sheet for fixture
    # economy; the extractor reads via named ranges so the sheet difference
    # is invisible.
    sub_factions = [
        # parent, sf_name, goal_axis, goal_delta, goal_text, infl,
        # m1, m2, m3, approval, national_share
        ("Founders", "Constitutional Loyalists", "authority", 1.0,
         "Defend the founding charter against revisionism.", 0.40,
         "", "", "", 0.5, 0.20),
        ("Founders", "Reformist Founders", "technocratic", 1.0,
         "Modernise the constitutional framework.", 0.35,
         "", "", "", 0.6, 0.18),
        ("Founders", "Hardliner Founders", "authority", 1.5,
         "Restore lost civic order through firm institutions.", 0.25,
         "", "", "", 0.4, 0.12),
        ("Capitalists", "Industrialists", "corporate", 1.5,
         "Expand heavy industry above all else.", 0.40,
         "", "", "", 0.5, 0.30),
        ("Capitalists", "Extraction Cartels", "expansion", 1.0,
         "Prioritise extraction over downstream value.", 0.35,
         "", "", "", 0.4, 0.20),
    ]
    # 6-axis effective stance per sub-faction (Expn, Auth, Corp, Tech, Faith,
    # Mat) — sourced from the live wb's Sub-Factions cols N:S. Distinct values
    # make it obvious in tests that the worldview comes straight from this
    # range and not from any baseline computation.
    sf_stances = {
        "Constitutional Loyalists": (4.0, 4.5, 4.0, 4.5, 5.0, 5.5),
        "Reformist Founders":       (3.5, 5.0, 4.0, 4.0, 4.5, 5.0),
        "Hardliner Founders":       (5.0, 6.2, 4.4, 4.1, 3.8, 3.5),
        "Industrialists":           (5.5, 3.5, 6.5, 4.0, 3.0, 2.5),
        "Extraction Cartels":       (6.5, 3.0, 6.5, 4.5, 3.5, 2.0),
    }
    for i, (parent, sf_name, goal_axis, goal_delta, goal_text, infl,
            m1, m2, m3, appr, nat_share) in enumerate(sub_factions, start=24):
        pol.cell(row=i, column=21, value=parent)       # U  (live col A)
        pol.cell(row=i, column=22, value=sf_name)      # V  (live col B)
        pol.cell(row=i, column=23, value=goal_axis)    # W  (live col C)
        pol.cell(row=i, column=24, value=goal_delta)   # X  (live col D)
        pol.cell(row=i, column=25, value=goal_text)    # Y  (live col E)
        pol.cell(row=i, column=26, value=infl)         # Z  (live col F)
        pol.cell(row=i, column=27, value=m1)           # AA (live col G)
        pol.cell(row=i, column=28, value=m2)           # AB (live col H)
        pol.cell(row=i, column=29, value=m3)           # AC (live col I)
        pol.cell(row=i, column=30, value=appr)         # AD (live col J)
        # Col AE (live col K) intentionally blank — live wb has Raw Nat. Weight
        # there, derived; no named range exposes it.
        pol.cell(row=i, column=32, value=nat_share)    # AF (live col L)
        stance = sf_stances.get(sf_name)
        if stance is not None:
            for k, v in enumerate(stance):
                pol.cell(row=i, column=33 + k, value=v)  # AG..AL (live col N..S)
    _add_name(wb, "SubFactionsBlock", "Politics!$U$24:$AF$36")
    _add_name(wb, "SubFactionGoals", "Politics!$U$24:$Y$36")        # A:E
    _add_name(wb, "SubFactionGoal", "Politics!$Y$24:$Y$36")          # E
    _add_name(wb, "SubFactionInfluences", "Politics!$Z$24:$Z$36")   # F
    _add_name(wb, "SubFactionMinorGoals", "Politics!$AA$24:$AC$36") # G:I
    _add_name(wb, "SubFactionApprovals", "Politics!$AD$24:$AD$36")  # J
    _add_name(wb, "SubFactionNationalShare", "Politics!$AF$24:$AF$36")  # L
    _add_name(wb, "SubFactionStances", "Politics!$AG$24:$AL$36")    # N:S

    # GoI Modifiers: PopCaptureBase B5:E15 (11 classes × 4 GoIs)
    gm = wb.create_sheet("GoI Modifiers")
    for i, (name, _, _, _) in enumerate(classes, start=5):
        gm.cell(row=i, column=1, value=name)
        for j in range(4):
            gm.cell(row=i, column=2 + j, value=0.20 + (j * 0.10))
    _add_name(wb, "PopCaptureBase", "'GoI Modifiers'!$B$5:$E$15")

    # GoI Benefits: A4:D15
    gb = wb.create_sheet("GoI Benefits")
    benefits = [
        ("Founders", 0.30, "Tax Holiday", "10% reduction"),
        ("Founders", 0.45, "Free Press", "Public approval +"),
        ("Founders", 0.60, "Constitutional Reform", "Stability +"),
        ("Capitalists", 0.30, "Subsidy", "Industry yield +"),
        ("Capitalists", 0.45, "Deregulation", "Crisis +"),
        ("Capitalists", 0.60, "Charter", "New corp"),
        ("Security", 0.30, "Patrol", "Security yield +"),
        ("Security", 0.45, "Curfew", "Stability + / approval -"),
        ("Security", 0.60, "Martial Law", "Big stab + / approval --"),
        ("Unionists", 0.30, "Min Wage", "Bargain +"),
        ("Unionists", 0.45, "Strike Right", "Bargain ++ / Crisis +"),
        ("Unionists", 0.60, "Co-op Charter", "New worker co-op"),
    ]
    for i, (goi, thresh, name, desc) in enumerate(benefits, start=4):
        gb.cell(row=i, column=1, value=goi)
        gb.cell(row=i, column=2, value=thresh)
        gb.cell(row=i, column=3, value=name)
        gb.cell(row=i, column=4, value=desc)
    _add_name(wb, "GoIBenefitsTable", "'GoI Benefits'!$A$4:$D$15")

    # Parties sheet: 15 slots rows 4-18. Two seeded for tests; the rest blank.
    pa = wb.create_sheet("Parties")
    pa["A1"] = "Parties Master"
    seeded = [
        # name, founded, establishment, 6-axis stance, weighted stance(6), Mad, ClosestGoI, Compat[4], ClassCompat[15], Estimated, VoteShare
        ("Liberty Now", True, 0.55, [5, 5, 4, 5, 4, 4]),
        ("People's Voice", True, 0.40, [3, 3, 6, 3, 4, 5]),
    ]
    for slot, (name, founded, est, stance) in enumerate(seeded):
        row = 4 + slot
        pa.cell(row=row, column=1, value=name)
        pa.cell(row=row, column=2, value=founded)
        pa.cell(row=row, column=3, value=est)
        for k, v in enumerate(stance):
            pa.cell(row=row, column=4 + k, value=v)
        # Weighted stance cols J-O (10-15) — same numbers for fixture
        for k, v in enumerate(stance):
            pa.cell(row=row, column=10 + k, value=v)
        pa.cell(row=row, column=16, value=0.15)              # Mad Index P
        pa.cell(row=row, column=17, value=["Founders", "Unionists"][slot])  # Closest GoI Q
        # GoI compat R-Y (cols 18-21 for 4 live GoIs)
        for j in range(4):
            pa.cell(row=row, column=18 + j, value=0.6 - j * 0.1 if slot == 0 else 0.3 + j * 0.1)
        # Class compat Z-AN (cols 26-40 for 15 class slots)
        for j in range(15):
            pa.cell(row=row, column=26 + j, value=0.5)
        pa.cell(row=row, column=41, value=0.30 if slot == 0 else 0.25)  # Estimated Support AO
        pa.cell(row=row, column=42, value=0.28 if slot == 0 else 0.22)  # Vote Share AP
    _add_name(wb, "PartiesBlock", "Parties!$A$4:$AP$18")

    # 11 map sheets, 40×40 each. Mostly empty terrain; a few seeded tiles.
    map_sheets = ["Terrain", "Features", "Resources", "Slots", "Improvements",
                  "Yield - Food", "Yield - Materials", "Yield - Ore",
                  "Yield - Energy", "Yield - Housing", "Yield - Water"]
    for sheet_name in map_sheets:
        ms = wb.create_sheet(sheet_name)
        for r in range(1, 41):
            for c in range(1, 41):
                if sheet_name == "Terrain":
                    ms.cell(row=r, column=c, value="Crater Floor" if (r + c) % 2 == 0 else "Mare Plain")
                elif sheet_name == "Slots":
                    ms.cell(row=r, column=c, value=2)
                elif sheet_name == "Features" and (r, c) == (5, 5):
                    ms.cell(row=r, column=c, value="Lava Tube")
                elif sheet_name == "Resources" and (r, c) == (10, 10):
                    ms.cell(row=r, column=c, value="He-3")
                elif sheet_name == "Improvements" and (r, c) == (10, 10):
                    ms.cell(row=r, column=c, value="HE3-1")
                elif sheet_name == "Yield - Energy" and (r, c) == (10, 10):
                    ms.cell(row=r, column=c, value=-1)
                else:
                    ms.cell(row=r, column=c, value=0 if sheet_name.startswith("Yield") else "")

    # Improvements manifest cols AO:AR (41-44) starting row 5.
    imp = wb["Improvements"]
    imp.cell(row=5, column=41, value="J10")           # Tile
    imp.cell(row=5, column=42, value="Helium-3 Mine")  # Improvement Type
    imp.cell(row=5, column=43, value="Corporate")      # Ownership Type
    imp.cell(row=5, column=44, value="Lunar Extractives")  # Owner

    # === New optional sheets for Map page (staffing + upkeep + workforce) ===
    # Spec: docs/superpowers/specs/2026-05-07-map-staffing-and-layer-dropdowns-design.md
    # All sheets are 40x40 sheet-keyed reads (no named ranges). Soft-fails on absence.

    # Staffing Efficiency: 40x40 grid of 0.0-1.0 floats. Seed a couple of cells.
    staffing_ws = wb.create_sheet("Staffing Efficiency")
    staffing_ws.cell(row=10, column=10, value=0.76)
    staffing_ws.cell(row=11, column=10, value=0.42)
    staffing_ws.cell(row=12, column=10, value=1.0)

    # Upkeep - <Resource>: 6 sheets of positive floats. Seed one cell per sheet.
    UPKEEP_RESOURCES = ["Food", "Water", "Energy", "Materials", "Ore", "Housing"]
    for resource in UPKEEP_RESOURCES:
        ws = wb.create_sheet(f"Upkeep - {resource}")
        ws.cell(row=10, column=10, value=1.5)

    # Workforce - <Class>: one sheet per class in ClassTable.
    # Names sourced from the fixture's ClassTable (rows 56-66 of Reference) — the spec
    # mandates ClassTable is the source of truth for workforce sheet names; do NOT
    # hardcode a parallel list that can drift.
    # Excel's 31-char sheet-name limit means class names >=20 chars produce
    # truncated sheet names — openpyxl truncates silently. Pre-truncating here
    # makes that explicit. The extractor in Task 5 looks up the FULL name from
    # ClassTable, so a truncated sheet naturally exercises the soft-fail path
    # for that class. Today only "Agricultural Workers" hits this in the fixture.
    workforce_classes = [name for (name, _tier, _pop, _weight) in classes]
    for cls in workforce_classes:
        sheet_name = f"Workforce - {cls}"[:31]
        ws = wb.create_sheet(sheet_name)
        ws.cell(row=10, column=10, value=12)  # 12 workers on tile (9, 9) — note 1-indexed
        ws.cell(row=11, column=10, value=0)   # explicit zero to validate drop-zero behaviour

    # Lookup helper for terrain → palette colour.
    pal = wb.create_sheet("MapPalette")
    palette = [
        ("Crater Floor", "#5a4a3a"),
        ("Mare Plain", "#3c3a3a"),
        ("Crater Rim", "#8a7560"),
        ("Polar Ice Plain", "#c8d8e8"),
        ("Empty", "#1a1a1a"),
    ]
    for i, (name, hex_) in enumerate(palette, start=1):
        pal.cell(row=i, column=1, value=name)
        pal.cell(row=i, column=2, value=hex_)
    _add_name(wb, "TerrainPalette", "MapPalette!$A$1:$B$10")

    # Coalitions sheet: 5 slots, rows 4-8.
    # Cols: A name, B-P (15) party-membership flags, then derived aggregates.
    co = wb.create_sheet("Coalitions")
    co["A1"] = "Coalitions"
    coalitions = [
        ("Big Tent", [True, True] + [False] * 13, 2, 0.95, 0.50, "Reformist"),
        ("Workers' Bloc", [False, True] + [False] * 13, 1, 0.40, 0.22, "Activist"),
    ]
    for slot, (name, flags, members, est, vote, approach) in enumerate(coalitions):
        row = 4 + slot
        co.cell(row=row, column=1, value=name)
        for j, flag in enumerate(flags):
            co.cell(row=row, column=2 + j, value=flag)
        # Derived cols: Q=member count, R=total establishment, S=total vote share,
        # T-Y=worldview centroid, Z=mad index, AA=approach
        co.cell(row=row, column=17, value=members)
        co.cell(row=row, column=18, value=est)
        co.cell(row=row, column=19, value=vote)
        for axis in range(6):
            co.cell(row=row, column=20 + axis, value=4.0 + slot * 0.5)
        co.cell(row=row, column=26, value=0.10)
        co.cell(row=row, column=27, value=approach)
    _add_name(wb, "CoalitionsBlock", "Coalitions!$A$4:$AA$8")

    # ---- ImprovementsCatalog sheet (soft-optional, v6) ----
    cat = wb.create_sheet("ImprovementsCatalog")
    headers = [
        "Name", "Category",
        "Mat Cost", "Ore Cost", "Eng Cost", "$ Cost",
        "Yield: Food", "Yield: Materials", "Yield: Ore", "Yield: Energy",
        "Yield: Housing", "Yield: Money", "Yield: Helium 3", "Yield: Water",
        "Yield: Stability", "Yield: Satisfaction All", "Yield: Research",
        "Upkeep: Energy", "Upkeep: Materials", "Upkeep: Money",
        "Upkeep: Ore", "Upkeep: Water",
        "Workforce: Bureaucrats", "Workforce: Capitalists", "Workforce: Engineers",
        "Workforce: Scientists", "Workforce: Security", "Workforce: Proprietors",
        "Workforce: Managerial", "Workforce: Botanists",
        "Workforce: Industrial Workers", "Workforce: Extraction Workers",
        "Workforce: Service Workers",
        "Split: Greens", "Split: Cereal", "Split: Vat Protein", "Split: Algal Paste",
        "Terrain Compatibility (notes)", "Ownership Options (notes)",
    ]
    for j, h in enumerate(headers, start=1):
        cat.cell(row=1, column=j, value=h)

    # Row 2: a normal Energy improvement (Solar Array Field).
    solar = ["Solar Array Field", "Energy",
             100, 0, 50, 200,
             0, 0, 0, 200, 0, 0, 0, 0, 0, 0, 0,
             0, 1, 0, 0, 0,
             0, 0, 2, 0, 0, 0, 0, 0, 0, 0, 0,
             None, None, None, None,
             "Mare Plain · Highlands", "Public · Private"]
    for j, v in enumerate(solar, start=1):
        cat.cell(row=2, column=j, value=v)

    # Row 3: blank-category Hydroponic Farm (forces categorizer regex fallback downstream).
    farm = ["Hydroponic Farm", None,
            150, 0, 100, 250,
            500, 0, 0, 0, 0, 0, 0, -50, 0, 0, 0,
            10, 5, 0, 0, 20,
            0, 0, 0, 0, 0, 0, 0, 4, 0, 0, 0,
            0.6, 0.0, 0.4, 0.0,
            "Mare Plain", "Public"]
    for j, v in enumerate(farm, start=1):
        cat.cell(row=3, column=j, value=v)

    # Row 4: unknown-category row (forces stdout warning + slug → 'other').
    weird = ["Heat Pump", "Power",
             50, 0, 25, 80,
             0, 0, 0, -10, 0, 0, 0, 0, 0, 0, 0,
             5, 0, 0, 0, 0,
             0, 0, 1, 0, 0, 0, 0, 0, 0, 0, 0,
             None, None, None, None,
             None, None]
    for j, v in enumerate(weird, start=1):
        cat.cell(row=4, column=j, value=v)

    # Row 5: blank-name (must be skipped by extractor — convention 8).

    _add_name(wb, "ImprovementsCatalog", "ImprovementsCatalog!$A$1:$AM$5")

    wb.save(out_path)
    return out_path


def _add_name(wb, name: str, attr_text: str) -> None:
    wb.defined_names[name] = DefinedName(name, attr_text=attr_text)


if __name__ == "__main__":
    import sys
    out = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("test_workbook.xlsx")
    build(out)
    print(f"Wrote {out}")
