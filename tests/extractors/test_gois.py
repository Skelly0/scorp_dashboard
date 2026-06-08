"""Tests for the GoIs page extractor."""
from __future__ import annotations

import openpyxl
from openpyxl.workbook.defined_name import DefinedName

from extractors.gois import (
    extract,
    _parse_active_benefits,
    _read_visible_goi_benefits_table,
)


def test_extract_returns_live_gois_only(wb):
    result = extract(wb)
    # Fixture has 5 live + 3 blank slots.
    assert [g["name"] for g in result["gois"]] == [
        "Administration",
        "Corporate",
        "Unions",
        "Security",
        "Research",
    ]


def test_extract_goi_record_shape(wb):
    g = extract(wb)["gois"][0]
    assert g["main_class"] == "Bureaucrats"
    assert g["derived_influence"] == 0.30
    assert g["approval"] == 0.55
    assert g["approach"] == "Reformist"
    assert isinstance(g["effective_worldview"], dict)
    assert g["mad_index"] == 0.10


def test_extract_active_benefits_parsed(wb):
    g = extract(wb)["gois"][0]
    assert g["active_benefits"]["unlocked"] == 1
    assert g["active_benefits"]["total"] == 3
    assert isinstance(g["active_benefits"]["unlocked_list"], list)


def test_extract_active_benefits_include_descriptions_and_status(wb):
    g = extract(wb)["gois"][0]

    assert g["active_benefits"]["items"] == [
        {
            "name": "Charter Draft",
            "description": "Stability +",
            "threshold": 0.30,
            "active": True,
        },
        {
            "name": "Civil Service",
            "description": "Admin capacity +",
            "threshold": 0.45,
            "active": False,
        },
        {
            "name": "Public Mandate",
            "description": "Approval +",
            "threshold": 0.60,
            "active": False,
        },
    ]
    assert g["active_benefits"]["unlocked_list"] == ["Charter Draft"]


def test_extract_benefits_falls_back_to_visible_table_when_named_range_stale(wb):
    wb.defined_names["GoIBenefitsTable"] = DefinedName(
        "GoIBenefitsTable",
        attr_text="'GoI Benefits'!$A$4:$D$15",
    )

    result = extract(wb)
    research = next(g for g in result["gois"] if g["name"] == "Research")

    assert [b["name"] for b in research["active_benefits"]["items"]] == [
        "Open Lab",
        "Peer Review",
        "Institute Charter",
    ]


def test_visible_benefits_table_tolerates_separator_rows():
    # A single blank spacer row between GoI groups must not truncate the table.
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GoI Benefits"
    ws["A1"] = "GOI BENEFITS"
    for col, header in enumerate(("GoI", "Threshold", "Benefit Name", "Description"), start=1):
        ws.cell(row=3, column=col, value=header)
    # rows 4-5: Administration; row 6 blank separator; row 7: Research
    data = {
        4: ("Administration", 0.30, "Charter Draft", "Stability +"),
        5: ("Administration", 0.45, "Civil Service", "Admin capacity +"),
        # row 6 intentionally blank
        7: ("Research", 0.30, "Open Lab", "Research yield +"),
    }
    for row, (goi, thresh, name, desc) in data.items():
        ws.cell(row=row, column=1, value=goi)
        ws.cell(row=row, column=2, value=thresh)
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=desc)

    rows = _read_visible_goi_benefits_table(wb)
    assert [r[2] for r in rows] == ["Charter Draft", "Civil Service", "Open Lab"]


def test_active_flags_follow_threshold_order_not_row_order():
    # Rows stored out of threshold order; the lowest-threshold benefits should
    # still be the active ones, matching the workbook's "2 / 3" count.
    table = [
        ["X", 0.60, "High", "d"],
        ["X", 0.30, "Low", "d"],
        ["X", 0.45, "Mid", "d"],
    ]
    result = _parse_active_benefits("2 / 3 unlocked", "X", table)

    assert [b["name"] for b in result["items"]] == ["Low", "Mid", "High"]
    assert [b["active"] for b in result["items"]] == [True, True, False]
    assert result["unlocked_list"] == ["Low", "Mid"]


def test_extract_sub_factions_grouped_under_parent(wb):
    result = extract(wb)
    security = next(g for g in result["gois"] if g["name"] == "Security")
    sf_names = [s["name"] for s in security["sub_factions"]]
    assert sf_names == ["Officers", "Fraternalists", "Shepherds"]


def test_extract_skips_structural_subfaction_rows_without_names(wb):
    ws = wb["Politics"]
    ws["V24"] = None
    ws["Y24"] = None
    ws["Z24"] = None
    ws["AD24"] = None
    ws["AF24"] = None
    for col in range(33, 39):
        ws.cell(row=24, column=col, value=None)

    result = extract(wb)
    all_sfs = [s for g in result["gois"] for s in g["sub_factions"]]

    assert all(s["name"] not in (None, "") for s in all_sfs)
    administration = next(g for g in result["gois"] if g["name"] == "Administration")
    assert [s["name"] for s in administration["sub_factions"]] == [
        "Reformist Administrators",
        "Hardliner Administrators",
    ]


def test_extract_prefers_visible_subfaction_detail_when_named_ranges_are_stale(wb):
    ws = wb["Politics"]
    live_row = 9
    ws.cell(row=live_row, column=1, value="Outcast Dissidents")
    ws.cell(row=live_row, column=6, value=0.31)
    ws.cell(row=live_row, column=8, value="0 / 0 unlocked")
    ws.cell(row=live_row, column=17, value=0.22)
    ws.cell(row=live_row, column=18, value="Dissident")
    ws.cell(row=live_row, column=19, value=0.09)
    for axis in range(6):
        ws.cell(row=live_row, column=11 + axis, value=4.0 + axis * 0.2)

    detail = wb.create_sheet("Sub-Faction Detail")
    headers = [
        "GoI",
        "Sub-faction",
        "Influence",
        "Goal Axis",
        "Goal Δ",
        "Expn",
        "Auth",
        "Corp",
        "Tech",
        "Faith",
        "Mat",
        "Approval",
        "Minor 1",
        "Minor 2",
        "Minor 3",
        "National Share",
    ]
    for col, header in enumerate(headers, start=1):
        detail.cell(row=3, column=col, value=header)
    rows = [
        ("Administration", "Legal Professionals", 0.5, None, 0, 4.45, 4.7, 3.45, 4.15, 4.3, 4.3, 0.5709, None, None, None, 0.0765),
        ("Administration", None, None, None, None, None, None, None, None, None, None, None, None, None, None, None),
        ("Outcast Dissidents", "Rebellious Youths", 0.1, None, 0, 4.0, 4.0, 5.4, 5.0, 3.55, 5.1, 0.5491, None, None, None, 0.0103),
        ("Outcast Dissidents", "Reactionaries", 0.5, None, 0, 3.4, 2.7, 1.0, 1.6, 3.3, 6.0, 0.2240, None, None, None, 0.0515),
    ]
    for row_idx, row in enumerate(rows, start=4):
        for col, value in enumerate(row, start=1):
            detail.cell(row=row_idx, column=col, value=value)

    result = extract(wb)
    outcasts = next(g for g in result["gois"] if g["name"] == "Outcast Dissidents")

    assert [s["name"] for s in outcasts["sub_factions"]] == [
        "Rebellious Youths",
        "Reactionaries",
    ]
    assert outcasts["sub_factions"][0]["influence"] == 0.1
    assert outcasts["sub_factions"][0]["approval"] == 0.5491
    assert outcasts["sub_factions"][0]["national_share"] == 0.0103
    assert outcasts["sub_factions"][0]["effective_worldview"] == {
        "expansion": 4.0,
        "authority": 4.0,
        "corporate": 5.4,
        "technocratic": 5.0,
        "faith": 3.55,
        "materialist": 5.1,
    }


def test_extract_pop_capture_matrix_shape(wb):
    matrix = extract(wb)["pop_capture_matrix"]
    assert len(matrix["classes"]) == 11
    assert len(matrix["gois"]) == 5
    assert len(matrix["values"]) == 11
    assert all(len(row) == 5 for row in matrix["values"])


def test_extract_pop_capture_matrix_uses_captured_pop_counts(wb):
    matrix = extract(wb)["pop_capture_matrix"]

    # The capture table's header order is Administration/Corporate/Unions/
    # Research/Security, while the Politics card order puts Security before
    # Research. Values should follow the displayed GoI labels, not raw columns.
    assert matrix["values"][0] == [1100, 1200, 1300, 1500, 1400]
    assert matrix["values"][-1] == [11100, 11200, 11300, 11500, 11400]


def test_extract_includes_subfaction_goal(wb):
    result = extract(wb)
    administration = next(g for g in result["gois"] if g["name"] == "Administration")
    statebuilders = next(s for s in administration["sub_factions"]
                         if s["name"] == "Statebuilders")
    assert statebuilders["goal"] == "Establish a constitution"


def test_extract_includes_subfaction_national_share(wb):
    result = extract(wb)
    security = next(g for g in result["gois"] if g["name"] == "Security")
    officers = next(s for s in security["sub_factions"]
                    if s["name"] == "Officers")
    assert officers["national_share"] == 0.05
    # All live sub-factions in the fixture should have a numeric national_share.
    all_sfs = [s for g in result["gois"] for s in g["sub_factions"]]
    assert all(isinstance(s["national_share"], float) for s in all_sfs)


def test_extract_includes_subfaction_effective_worldview(wb):
    """Per-axis effective worldview is read directly from SubFactionStances
    (Sub-Factions!N:S in the live wb; mirrored at Politics!AG:AL in the
    fixture, row-aligned with SubFactionGoals)."""
    result = extract(wb)
    security = next(g for g in result["gois"] if g["name"] == "Security")
    officers = next(s for s in security["sub_factions"]
                    if s["name"] == "Officers")
    ew = officers["effective_worldview"]
    assert isinstance(ew, dict)
    # Officers' fixture stance row, copied verbatim from the
    # SubFactionStances range — no baseline-plus-delta math involved.
    assert ew == {
        "expansion": 5.0,
        "authority": 1.65,
        "corporate": 4.0,
        "technocratic": 3.1,
        "faith": 3.5,
        "materialist": 2.5,
    }


def test_extract_handles_missing_subfaction_ranges(fixture_workbook_path):
    """When soft-optional sub-faction ranges are absent, extraction still
    works — the missing fields default to None. SubFactionStances remains
    so the radar still has data; SubFactionGoal/NationalShare drop to —."""
    wb = openpyxl.load_workbook(fixture_workbook_path, data_only=True)
    # Remove the soft-optional text/share ranges to simulate an older workbook.
    for nm in ("SubFactionGoal", "SubFactionNationalShare"):
        if nm in wb.defined_names:
            del wb.defined_names[nm]

    result = extract(wb)
    all_sfs = [s for g in result["gois"] for s in g["sub_factions"]]
    # Fixture has 15 live sub-factions across 5 GoIs.
    assert len(all_sfs) == 15, "removing names should not drop sub-factions"
    for sf in all_sfs:
        assert sf["goal"] is None
        assert sf["national_share"] is None
        # Worldview still populated — SubFactionStances is still present.
        assert isinstance(sf["effective_worldview"], dict)


def test_extract_handles_missing_subfaction_stances(fixture_workbook_path):
    """When SubFactionStances itself is absent, the radar is hidden — every
    sub-faction's effective_worldview becomes None (graceful degradation)."""
    wb = openpyxl.load_workbook(fixture_workbook_path, data_only=True)
    if "SubFactionStances" in wb.defined_names:
        del wb.defined_names["SubFactionStances"]

    result = extract(wb)
    all_sfs = [s for g in result["gois"] for s in g["sub_factions"]]
    assert len(all_sfs) == 15, "removing the stance range must not drop sub-factions"
    for sf in all_sfs:
        assert sf["effective_worldview"] is None
