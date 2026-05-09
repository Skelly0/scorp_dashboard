"""Tests for the Status page extractor."""
from __future__ import annotations

import pytest
from openpyxl.workbook.defined_name import DefinedName

from extractors.status import extract


def set_name(wb, name: str, attr_text: str) -> None:
    wb.defined_names[name] = DefinedName(name, attr_text=attr_text)


def test_extract_returns_treasury_block(wb):
    result = extract(wb)
    assert result["treasury"] == {"money": 487, "delta": -12}


def test_extract_treasury_fills_partial_legacy_names_from_live_layout(wb):
    del wb.defined_names["TreasuryMoneyDelta"]
    col = wb["Colony"]
    col["B3"] = 900
    col["A8"], col["B8"], col["C8"], col["D8"] = "Money", 900, 40, 10

    result = extract(wb)

    assert result["treasury"] == {"money": 487, "delta": 30}


def test_extract_returns_stability_and_crisis(wb):
    result = extract(wb)
    assert result["stability"] == 0.42
    assert result["crisis_factor"] == 0.38


def test_extract_returns_population_total(wb):
    result = extract(wb)
    # Sum of fixture's class populations.
    assert result["population_total"] == 15870


def test_extract_returns_resource_flow_strip(wb):
    result = extract(wb)
    resources = result["resources"]
    names = [r["name"] for r in resources]
    assert names == ["Food", "Materials", "Ore", "Energy", "Housing", "He-3", "Water"]
    assert resources[0] == {"name": "Food", "current": 0, "delta": -2}
    assert resources[6] == {"name": "Water", "current": 60, "delta": -1}


def test_extract_resource_flow_accepts_live_four_column_shape(wb):
    col = wb["Colony"]
    col.cell(row=4, column=3, value=5)
    col.cell(row=4, column=4, value=7)
    set_name(wb, "ResourceFlows", "Colony!$A$4:$D$10")

    result = extract(wb)

    assert result["resources"][0] == {
        "name": "Food",
        "current": 0,
        "income": 5,
        "upkeep": 7,
        "delta": -2,
    }


def test_extract_resource_flow_accepts_live_six_column_shape(wb):
    col = wb["Colony"]
    col.cell(row=4, column=3, value=20)
    col.cell(row=4, column=4, value=7)
    col.cell(row=4, column=5, value=3)
    col.cell(row=4, column=6, value=10)
    set_name(wb, "ResourceFlows", "Colony!$A$4:$F$10")

    result = extract(wb)

    assert result["resources"][0] == {
        "name": "Food",
        "current": 0,
        "income": 20,
        "upkeep": 10,
        "delta": 10,
    }


def test_extract_resource_flow_falls_back_when_named_range_is_blank(wb):
    col = wb["Colony"]
    for row in range(8, 16):
        for column in range(1, 7):
            col.cell(row=row, column=column).value = None
    col["A8"], col["B8"], col["C8"], col["D8"], col["E8"], col["F8"] = "Money", 900, 40, 10, 2, 28
    col["A9"], col["B9"], col["C9"], col["D9"], col["E9"], col["F9"] = "Food", 100, 5, 7, 3, -5
    set_name(wb, "ResourceFlows", "Colony!$Z$1:$AB$3")

    result = extract(wb)

    assert result["resources"] == [
        {"name": "Money", "current": 900, "income": 40, "upkeep": 12, "delta": 28},
        {"name": "Food", "current": 100, "income": 5, "upkeep": 10, "delta": -5},
    ]


def test_extract_returns_overton_window(wb):
    result = extract(wb)
    assert result["overton"] == {
        "expansion": 5.0,
        "authority": 4.5,
        "corporate": 5.0,
        "technocratic": 4.0,
        "faith": 4.0,
        "materialist": 4.0,
    }


def test_extract_active_situations_empty_when_no_situations_sheet(wb):
    """Wave 1: Situations sheet doesn't exist yet — extractor should return [] not crash."""
    result = extract(wb)
    assert result["active_situations"] == []


def test_extract_returns_year(wb):
    result = extract(wb)
    assert result["year"] == 12


def test_extract_year_none_when_named_range_missing(wb):
    del wb.defined_names["Var_Year"]
    result = extract(wb)
    assert result["year"] is None


def test_extract_returns_gov_approval(wb):
    result = extract(wb)
    assert result["gov_approval"] == 0.62


def test_extract_gov_approval_none_when_range_missing(wb):
    del wb.defined_names["EffectiveGovApproval"]
    result = extract(wb)
    assert result["gov_approval"] is None


def test_avg_satisfaction_population_weighted(wb):
    """Weighted mean across PopsimSatisfaction × PopsimPop."""
    from extractors._common import avg_satisfaction
    # Fixture: all sat = 0.40, all pops varied. Weighted mean = 0.40.
    assert avg_satisfaction(wb) == 0.40


def test_avg_satisfaction_returns_none_when_total_pop_zero(wb):
    """Guard against division by zero when every pop cell is zero/None."""
    from extractors._common import avg_satisfaction
    pop_sheet = wb["Popsim"]
    for row in range(5, 20):  # PopsimPop range B5:B19
        pop_sheet.cell(row=row, column=2, value=0)
    assert avg_satisfaction(wb) is None


def test_housing_util_normal_case():
    """pop / capacity returns ratio in 0..∞ range."""
    from extractors.status import _housing_util
    assert _housing_util(15870, 16500) == 15870 / 16500


def test_housing_util_returns_none_when_capacity_zero():
    """Div-by-zero guard."""
    from extractors.status import _housing_util
    assert _housing_util(15870, 0) is None
    assert _housing_util(15870, None) is None
    assert _housing_util(0, 16500) == 0.0  # zero pop OK; zero capacity not.


def test_net_delta_pct_normal():
    from extractors._common import net_delta_pct
    # growth=0.02, cdr=0.012 → (0.02 - 0.012) * 100 = 0.8
    assert net_delta_pct(0.020, 0.012) == pytest.approx(0.8)


def test_net_delta_pct_none_when_either_input_missing():
    from extractors._common import net_delta_pct
    assert net_delta_pct(None, 0.012) is None
    assert net_delta_pct(0.020, None) is None
    assert net_delta_pct(None, None) is None


def test_extract_demographics_block_shape(wb):
    result = extract(wb)
    demo = result["demographics"]
    assert set(demo.keys()) == {
        "base_growth_rate", "sat_elasticity", "effective_growth_rate",
        "effective_cdr", "total_births", "total_deaths", "net_delta_pct",
        "housing_capacity", "housing_util", "avg_satisfaction",
    }


def test_extract_demographics_values_from_fixture(wb):
    result = extract(wb)
    demo = result["demographics"]
    assert demo["base_growth_rate"] == 0.020
    assert demo["sat_elasticity"] == 0.95
    assert demo["effective_growth_rate"] == pytest.approx(0.020 * 0.95)
    assert demo["effective_cdr"] == 0.0125
    assert demo["total_births"] == 320
    assert demo["total_deaths"] == 280
    assert demo["housing_capacity"] == 16500
    assert demo["avg_satisfaction"] == 0.40


def test_extract_demographics_total_births_soft_optional(wb):
    del wb.defined_names["TotalBirths"]
    result = extract(wb)
    demo = result["demographics"]
    assert demo["total_births"] is None
    assert demo["total_deaths"] == 280


def test_extract_demographics_effective_growth_none_when_base_missing(wb):
    del wb.defined_names["Var_BaseGrowthRate"]
    result = extract(wb)
    demo = result["demographics"]
    assert demo["base_growth_rate"] is None
    assert demo["effective_growth_rate"] is None
    assert demo["net_delta_pct"] is None  # chains through
