"""Tests for the ImprovementsCatalog extractor."""
from __future__ import annotations

import logging

from extractors.catalog import extract


def test_extract_returns_improvements_list(wb):
    result = extract(wb)
    assert "improvements" in result
    assert isinstance(result["improvements"], list)


def test_extract_skips_blank_name_row(wb):
    result = extract(wb)
    names = [imp["name"] for imp in result["improvements"]]
    assert "Solar Array Field" in names
    assert "Hydroponic Farm" in names
    assert "Heat Pump" in names
    assert None not in names
    assert "" not in names
    assert len(names) == 3


def test_extract_solar_field_row(wb):
    result = extract(wb)
    solar = next(i for i in result["improvements"] if i["name"] == "Solar Array Field")
    assert solar["category"] == "energy"
    assert solar["category_raw"] == "Energy"
    assert solar["costs"] == {"materials": 100, "ore": 0, "engineering": 50, "money": 200}
    assert solar["yields"]["energy"] == 200
    assert solar["yields"]["food"] == 0
    assert solar["upkeep"]["materials"] == 1
    assert solar["workforce"]["engineers"] == 2
    assert solar["splits"] == {"greens": None, "cereal": None,
                               "vat_protein": None, "algal_paste": None}
    assert solar["terrain_compat"] == "Mare Plain · Highlands"
    assert solar["ownership_options"] == "Public · Private"


def test_extract_blank_category_falls_through(wb):
    result = extract(wb)
    farm = next(i for i in result["improvements"] if i["name"] == "Hydroponic Farm")
    assert farm["category"] is None
    assert farm["category_raw"] is None
    assert farm["splits"]["greens"] == 0.6
    assert farm["splits"]["vat_protein"] == 0.4
    assert farm["yields"]["food"] == 500
    assert farm["yields"]["water"] == -50


def test_extract_unknown_category_logs_warning_and_maps_to_other(wb, caplog):
    with caplog.at_level(logging.WARNING, logger="extractors.catalog"):
        result = extract(wb)
    pump = next(i for i in result["improvements"] if i["name"] == "Heat Pump")
    assert pump["category"] == "other"
    assert pump["category_raw"] == "Power"
    assert any("Power" in r.message and "Heat Pump" in r.message
               for r in caplog.records)


def test_extract_zeros_preserved_blanks_coerced_to_zero(wb):
    """Numeric blanks coerce to 0 (per spec) — not None."""
    result = extract(wb)
    solar = next(i for i in result["improvements"] if i["name"] == "Solar Array Field")
    assert solar["yields"]["food"] == 0
    assert solar["yields"]["materials"] == 0
    assert solar["workforce"]["bureaucrats"] == 0


def test_extract_with_missing_range_returns_empty():
    """If ImprovementsCatalog isn't defined, extract returns empty list."""
    import openpyxl
    blank_wb = openpyxl.Workbook()
    result = extract(blank_wb)
    assert result == {"improvements": []}


def test_extract_live_shape_when_named_range_starts_below_headers(wb):
    """Live workbook range starts at its first data row; headers sit one row above.

    Arithmetic mirrors the live `ImprovementsCatalog!$A$4:$AO$44` contract
    (data row 4, header on row 3); here the 3-row fixture is shifted by one
    row so the same one-row offset is exercised.
    """
    from openpyxl.workbook.defined_name import DefinedName

    wb.defined_names["ImprovementsCatalog"] = DefinedName(
        "ImprovementsCatalog",
        attr_text="ImprovementsCatalog!$A$2:$AM$5",
    )

    result = extract(wb)

    names = [imp["name"] for imp in result["improvements"]]
    assert names == ["Solar Array Field", "Hydroponic Farm", "Heat Pump"]
    assert result["improvements"][0]["category"] == "energy"
    assert result["improvements"][1]["splits"]["greens"] == 0.6


def test_extract_warns_when_named_range_pinned_to_row_one_without_header(caplog):
    """If `ImprovementsCatalog` starts at row 1 and that row isn't a header,
    there is no row above to read headers from; the extractor warns and bails
    to an empty list instead of misreading data rows as headers silently.
    """
    import openpyxl
    from openpyxl.workbook.defined_name import DefinedName

    bw = openpyxl.Workbook()
    ws = bw.active
    ws.title = "ImprovementsCatalog"
    ws["A1"], ws["B1"] = "Hydroponic Farm", "Food"
    ws["A2"], ws["B2"] = "Solar Array", "Energy"
    bw.defined_names["ImprovementsCatalog"] = DefinedName(
        "ImprovementsCatalog",
        attr_text="ImprovementsCatalog!$A$1:$AM$2",
    )

    with caplog.at_level(logging.WARNING, logger="extractors.catalog"):
        result = extract(bw)

    assert result == {"improvements": []}
    assert any(
        "ImprovementsCatalog" in r.message and "row 1" in r.message
        for r in caplog.records
    )


def test_extract_tolerates_whitespace_around_slashes_in_category(wb, caplog):
    """`Faith / Culture / Recreation` and `Faith/Culture/Recreation` map to
    the same dashboard slug without warning."""
    from openpyxl.workbook.defined_name import DefinedName

    ws = wb["ImprovementsCatalog"]
    ws["A2"], ws["B2"] = "Faith Center", "Faith / Culture / Recreation"
    ws["A3"], ws["B3"] = "Research Lab", "Research / Economy"
    wb.defined_names["ImprovementsCatalog"] = DefinedName(
        "ImprovementsCatalog",
        attr_text="ImprovementsCatalog!$A$1:$AM$3",
    )

    with caplog.at_level(logging.WARNING, logger="extractors.catalog"):
        result = extract(wb)

    by_name = {imp["name"]: imp for imp in result["improvements"]}
    assert by_name["Faith Center"]["category"] == "civic"
    assert by_name["Research Lab"]["category"] == "science"
    assert caplog.records == []


def test_extract_accepts_live_category_vocabulary(wb, caplog):
    """Live workbook category labels map into dashboard slugs without warnings."""
    from openpyxl.workbook.defined_name import DefinedName

    ws = wb["ImprovementsCatalog"]
    ws["A2"], ws["B2"] = "Hydroponic Bay", "Food"
    ws["A3"], ws["B3"] = "Materials Refinery", "Materials"
    ws["A4"], ws["B4"] = "Surface Hab Module", "Housing"
    ws["A5"], ws["B5"] = "Faith Center", "Faith/Culture/Recreation"
    ws["A6"], ws["B6"] = "Research Lab", "Research/Economy"
    wb.defined_names["ImprovementsCatalog"] = DefinedName(
        "ImprovementsCatalog",
        attr_text="ImprovementsCatalog!$A$1:$AM$6",
    )

    with caplog.at_level(logging.WARNING, logger="extractors.catalog"):
        result = extract(wb)

    by_name = {imp["name"]: imp for imp in result["improvements"]}
    assert by_name["Hydroponic Bay"]["category"] == "agri"
    assert by_name["Materials Refinery"]["category"] == "materials"
    assert by_name["Surface Hab Module"]["category"] == "habitat"
    assert by_name["Faith Center"]["category"] == "civic"
    assert by_name["Research Lab"]["category"] == "science"
    assert caplog.records == []
