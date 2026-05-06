"""Tests for the Situations page extractor."""
from __future__ import annotations

import openpyxl
import pytest

from extractors.situations import extract


def test_extract_returns_empty_when_sheets_missing(wb):
    result = extract(wb)
    assert result == {
        "active": [],
        "ended": [],
        "stability_modifiers": [],
        "tier_ladder": [],
    }


def test_extract_reads_situations_sheet_when_present(wb):
    ws = wb.create_sheet("Situations")
    # Live layout (CLAUDE.md gotcha #13): banner row 1, header row 2, data row 3+.
    # Columns: Name | Crisis Contribution | Description | Tier.
    ws["A1"] = "Active Situations"
    ws["A2"], ws["B2"], ws["C2"], ws["D2"] = "Name", "Crisis Contribution", "Description", "Tier"
    ws["A3"], ws["B3"], ws["C3"], ws["D3"] = "The Fall", 0.05, "Earth has fallen", "T2"
    ws["A4"], ws["B4"], ws["C4"], ws["D4"] = "Old Crisis", "Ended", "Resolved already", None
    result = extract(wb)
    assert len(result["active"]) == 1
    assert result["active"][0]["name"] == "The Fall"
    assert result["active"][0]["crisis_factor"] == 0.05
    assert result["active"][0]["description"] == "Earth has fallen"
    assert len(result["ended"]) == 1
    assert result["ended"][0]["name"] == "Old Crisis"


def test_extract_reads_stability_modifiers(wb):
    ws = wb.create_sheet("Stability Modifiers")
    ws["A1"], ws["B1"], ws["C1"] = "Name", "Description", "Factor"
    ws["A2"], ws["B2"], ws["C2"] = "Cohesive Founders", "Strong norms", 0.05
    ws["A3"], ws["B3"], ws["C3"] = "Worker Unrest", "Striking", -0.10
    result = extract(wb)
    assert len(result["stability_modifiers"]) == 2
    assert result["stability_modifiers"][1]["factor"] == -0.10


def test_extract_reads_tier_ladder(wb):
    ws = wb.create_sheet("Tier Ladder")
    ws["A1"], ws["B1"], ws["C1"] = "Tier", "Active", "Consequence"
    rows = [("I", False, "Civil unrest"), ("II", True, "Riots"), ("III", False, "Strikes")]
    for i, (t, a, c) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=t)
        ws.cell(row=i, column=2, value=a)
        ws.cell(row=i, column=3, value=c)
    result = extract(wb)
    assert len(result["tier_ladder"]) == 3
    assert result["tier_ladder"][1]["active"] is True
    assert result["tier_ladder"][1]["tier"] == "II"
