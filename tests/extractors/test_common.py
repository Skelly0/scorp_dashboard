"""Tests for extractor common helpers."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest

from extractors._common import (
    coerce_number,
    filter_blank_rows,
    read_named_range,
)


@pytest.fixture
def wb():
    """Tiny workbook with one named range and a blank-slot row."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    # A1:C3 grid with a blank middle row.
    ws["A1"], ws["B1"], ws["C1"] = "Alpha", 1.5, "x"
    ws["A2"], ws["B2"], ws["C2"] = "", "", ""
    ws["A3"], ws["B3"], ws["C3"] = "Bravo", 2.5, "y"
    dn = openpyxl.workbook.defined_name.DefinedName("MyRange", attr_text="Sheet1!$A$1:$C$3")
    wb.defined_names["MyRange"] = dn
    return wb


def test_read_named_range_returns_rows(wb):
    rows = read_named_range(wb, "MyRange")
    assert rows == [
        ["Alpha", 1.5, "x"],
        ["", "", ""],
        ["Bravo", 2.5, "y"],
    ]


def test_read_named_range_missing_returns_empty(wb):
    assert read_named_range(wb, "Nonexistent") == []


def test_filter_blank_rows_removes_empty_first_col(wb):
    rows = read_named_range(wb, "MyRange")
    filtered = filter_blank_rows(rows)
    assert filtered == [
        ["Alpha", 1.5, "x"],
        ["Bravo", 2.5, "y"],
    ]


def test_coerce_number_handles_floats():
    assert coerce_number(1.5) == 1.5
    assert coerce_number(0) == 0.0


def test_coerce_number_returns_none_for_blanks():
    assert coerce_number(None) is None
    assert coerce_number("") is None


def test_coerce_number_returns_none_for_formula_errors():
    assert coerce_number("#REF!") is None
    assert coerce_number("#NAME?") is None
    assert coerce_number("#VALUE!") is None


def test_fixture_workbook_loads(fixture_workbook_path):
    """Sanity: fixture loads with expected named ranges."""
    wb = openpyxl.load_workbook(fixture_workbook_path, data_only=True)
    assert "ClassTable" in wb.defined_names
    assert "Var_SenatePageVisible" in wb.defined_names
    assert "PopsimPop" in wb.defined_names
