"""Tests for the Map page extractor and the soft-fail helper in _common."""
from __future__ import annotations

import openpyxl
import pytest

from extractors._common import read_grid_optional
from extractors.map import extract


def _wb_with(sheet_name: str, values: list[list]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    # remove the default sheet so the workbook only has what the test asks for
    default = wb.active
    wb.remove(default)
    ws = wb.create_sheet(sheet_name)
    for r_idx, row in enumerate(values, start=1):
        for c_idx, v in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=v)
    return wb


def test_read_grid_optional_returns_grid_when_sheet_exists():
    wb = _wb_with("Staffing Efficiency", [[0.5, 0.7], [0.0, 1.0]])
    grid = read_grid_optional(wb, "Staffing Efficiency", width=2, height=2)
    assert grid == [[0.5, 0.7], [0.0, 1.0]]


def test_read_grid_optional_returns_none_when_sheet_missing():
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("Some Other Sheet")
    grid = read_grid_optional(wb, "Staffing Efficiency", width=2, height=2)
    assert grid is None


def test_read_grid_optional_pads_short_rows_with_none():
    """A sheet that's smaller than width x height fills missing cells with None."""
    wb = _wb_with("X", [[1, 2], [3]])  # second row has only 1 value
    grid = read_grid_optional(wb, "X", width=3, height=2)
    assert grid == [[1, 2, None], [3, None, None]]


def test_extract_dimensions(wb):
    result = extract(wb)
    assert result["width"] == 40
    assert result["height"] == 40
    assert len(result["tiles"]) == 1600


def test_extract_tile_record_shape(wb):
    result = extract(wb)
    # Tile (0, 0) — top-left corner
    t = result["tiles"][0]
    assert t["x"] == 0
    assert t["y"] == 0
    assert t["terrain"] in {"Crater Floor", "Mare Plain"}
    assert t["feature"] is None
    assert t["resource"] is None
    assert t["slots"] == 2
    assert t["improvement"] is None
    assert "yields" in t
    assert set(t["yields"].keys()) == {"food", "materials", "ore", "energy", "housing", "water"}


def test_extract_seeded_he3_tile(wb):
    """Tile at (9, 9) — 0-indexed coords for sheet row 10, col 10."""
    result = extract(wb)
    tile = next(t for t in result["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert tile["resource"] == "He-3"
    assert tile["improvement"] is not None
    assert tile["improvement"]["name"] == "Helium-3 Mine"
    assert tile["improvement"]["owner"] == "Lunar Extractives"
    assert tile["improvement"]["ownership_type"] == "Corporate"
    assert tile["control"] == "Founders"
    assert tile["yields"]["energy"] == -1


def test_extract_includes_palette(wb):
    result = extract(wb)
    assert "palettes" in result
    assert "terrain" in result["palettes"]
    assert result["palettes"]["terrain"]["Crater Floor"] == "#5a4a3a"


def test_extract_includes_resource_palette(wb):
    result = extract(wb)
    assert "resource" in result["palettes"]
    pal = result["palettes"]["resource"]
    # Must cover every resource type the live workbook can emit.
    assert pal["Helium-3"] == "#ffd166"
    assert pal["Iron Deposit"] == "#c97064"
    assert pal["Aluminum Deposit"] == "#b8c5d6"
    assert pal["Phosphorus Deposit"] == "#d6a8e0"
    assert pal["Rare Earths"] == "#7ed4a8"
    assert pal["Heavy Metals"] == "#6a7e9c"
    assert pal["Oxygen Bound Soil"] == "#5fc3e8"
    assert pal["Water Ice"] == "#ffffff"


def test_extract_includes_feature_palette(wb):
    result = extract(wb)
    assert "feature" in result["palettes"]
    pal = result["palettes"]["feature"]
    assert pal["Buried Ice"] == "#b3d9ff"
    assert pal["Mineral Vein"] == "#c4a484"
    assert pal["Smooth Plain"] == "#8a9da6"
    assert pal["Boulder Field"] == "#6e6058"
    assert pal["Cave System"] == "#2d3a4a"
    assert pal["Recent Meteorite Strikes"] == "#d97a5b"
    assert pal["Magnetic Anomaly"] == "#a89cff"
    assert pal["Hollow Rocks"] == "#a89567"
    assert pal["Crashed Probe"] == "#ff8c42"


def test_extract_includes_improvement_category_palette(wb):
    result = extract(wb)
    assert "improvement_category" in result["palettes"]
    pal = result["palettes"]["improvement_category"]
    assert pal["energy"] == "#ffb000"
    assert pal["mining"] == "#a06840"
    assert pal["habitat"] == "#7ed4a8"
    assert pal["civic"] == "#5ec3ff"
    assert pal["military"] == "#ff5544"
    assert pal["agri"] == "#38d39f"
    assert pal["science"] == "#a89cff"
    assert pal["other"] == "#888888"


def test_extract_includes_control_palette(wb):
    result = extract(wb)
    assert "control" in result["palettes"]
    pal = result["palettes"]["control"]
    assert pal["Administration"] == "#5ec3ff"
    assert pal["Corporations"] == "#ffd84d"
    assert pal["Founders"] == "#ffb000"


def test_staffing_present_when_sheet_exists(wb):
    out = extract(wb)
    assert out["available_categories"]["staffing"] is True
    # Tile (9, 9) was seeded with 0.76 in the fixture (1-indexed row 10, col 10)
    tile = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert tile["staffing"] == pytest.approx(0.76)


def test_staffing_absent_when_sheet_missing(wb):
    wb.remove(wb["Staffing Efficiency"])
    out = extract(wb)
    assert out["available_categories"]["staffing"] is False
    for tile in out["tiles"]:
        assert tile["staffing"] is None


def test_upkeep_present_when_all_sheets_exist(wb):
    out = extract(wb)
    assert out["available_categories"]["upkeep"] is True
    tile = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert tile["upkeep"]["food"] == pytest.approx(1.5)
    assert tile["upkeep"]["water"] == pytest.approx(1.5)


def test_upkeep_reads_live_sheet_names_without_dash(wb):
    for s in [name for name in wb.sheetnames if name.startswith("Upkeep - ")]:
        wb.remove(wb[s])
    ws = wb.create_sheet("Upkeep Energy")
    ws.cell(row=10, column=10, value=7.5)

    out = extract(wb)

    assert out["available_categories"]["upkeep"] is True
    tile = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert tile["upkeep"]["energy"] == pytest.approx(7.5)


def test_upkeep_absent_when_all_sheets_missing(wb):
    for r in ["Food", "Water", "Energy", "Materials", "Ore", "Housing"]:
        wb.remove(wb[f"Upkeep - {r}"])
    out = extract(wb)
    assert out["available_categories"]["upkeep"] is False
    for tile in out["tiles"]:
        assert tile["upkeep"] is None


def test_upkeep_partially_populated_when_some_sheets_missing(wb):
    """Mixed: 4 of 6 upkeep sheets present → upkeep is a dict with only those 4 keys."""
    wb.remove(wb["Upkeep - Ore"])
    wb.remove(wb["Upkeep - Housing"])
    out = extract(wb)
    assert out["available_categories"]["upkeep"] is True
    tile = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert "ore" not in tile["upkeep"]
    assert "housing" not in tile["upkeep"]
    assert "food" in tile["upkeep"]


def test_workforce_present_when_classtable_and_sheets_exist(wb):
    out = extract(wb)
    assert out["available_categories"]["workforce"] is True
    tile = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert tile["workforce"] is not None
    assert tile["workforce"].get("Engineers") == 12
    # zero entries are dropped — value at (9, 10) seeded with 0 should not appear
    tile_zero = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 10)
    assert "Engineers" not in (tile_zero["workforce"] or {})


def test_workforce_reads_live_sheet_names_without_dash(wb):
    for s in [name for name in wb.sheetnames if name.startswith("Workforce - ")]:
        wb.remove(wb[s])
    ws = wb.create_sheet("Workforce Engineers")
    ws.cell(row=10, column=10, value=44)

    out = extract(wb)

    assert out["available_categories"]["workforce"] is True
    tile = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert tile["workforce"]["Engineers"] == 44


def test_workforce_absent_when_all_workforce_sheets_missing(wb):
    workforce_sheets = [s for s in wb.sheetnames if s.startswith("Workforce - ")]
    for s in workforce_sheets:
        wb.remove(wb[s])
    out = extract(wb)
    assert out["available_categories"]["workforce"] is False
    for tile in out["tiles"]:
        assert tile["workforce"] is None


def test_workforce_skips_unknown_class_sheets(wb):
    """A workforce sheet for a class NOT in ClassTable is silently ignored."""
    wb.create_sheet("Workforce - Phantom Class")
    out = extract(wb)
    tile = next(t for t in out["tiles"] if t["x"] == 9 and t["y"] == 9)
    assert "Phantom Class" not in (tile["workforce"] or {})


def test_missing_sheets_reported_in_map_output(wb):
    wb.remove(wb["Staffing Efficiency"])
    wb.remove(wb["Upkeep - Ore"])
    out = extract(wb)
    sheets_missed = {f["sheet"] for f in out["missing_sheets"]}
    assert "Staffing Efficiency" in sheets_missed
    assert "Upkeep Ore" in sheets_missed
    for f in out["missing_sheets"]:
        assert f["kind"] == "missing_sheet"


def test_missing_sheets_includes_truncated_workforce_class(wb):
    """The fixture pre-truncates one legacy Workforce sheet to 31 chars; extractor
    looks for the live/full names and should record it as missing."""
    out = extract(wb)
    missing = {f["sheet"] for f in out["missing_sheets"]}
    assert "Workforce Agricultural Workers" in missing
